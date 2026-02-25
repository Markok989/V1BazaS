# FILENAME: ui/asset_detail_dialog.py
# (FILENAME: ui/asset_detail_dialog.py - START PART 1/4)
# -*- coding: utf-8 -*-
"""
BazaS2 (offline) — ui/asset_detail_dialog.py

Asset Detail Dialog:
- Tabovi: Detalji / Timeline / Kalendar / Metrologija / Dodatna polja / Prilozi
- Stabilan UI state (QSettings): geometry + splitteri + tab index
- Fail-safe importi (metrology/calendar/edit)
- Prilozi: preview (slike + tekstualni best-effort) + pretraga + open/open-folder + copy path
- Offline-only

Senior rev (fix):
- FIX: vraćen kompletan attachments blok + context menu slot (_on_tbl_context_menu),
  da UI ne puca na connect().
- SECURITY: asset read preferira services.assets_service.get_asset_by_uid (RBAC + scope),
  a direct DB read je fallback.
- Timeline: prefer core.db.list_asset_events_db, fallback legacy scan.
- Disposal: smart meni (best-effort), bez rušenja ako nema disposal tabela/funkcija.
"""

from __future__ import annotations

import logging
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from PySide6.QtCore import (
    Qt,
    QUrl,
    QSettings,
    QTimer,
    QByteArray,
    QObject,
    Signal,
    QThread,
)  # type: ignore
from PySide6.QtGui import QDesktopServices, QPixmap, QTextCursor, QTextDocument  # type: ignore
from PySide6.QtWidgets import (  # type: ignore
    QApplication,
    QAbstractItemView,
    QCalendarWidget,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QInputDialog,
)

from core.config import DB_FILE

# ✅ Session helpers (fail-safe)
try:
    from core.session import actor_name as actor_name, can as can, current_sector as current_sector  # type: ignore
except Exception:  # pragma: no cover
    def actor_name() -> str:  # type: ignore
        return "user"

    def can(_perm: str) -> bool:  # type: ignore
        return False

    def current_sector() -> str:  # type: ignore
        return ""


from ui.utils.datetime_fmt import fmt_date_sr

from services.attachments_service import (
    list_attachments_for_asset,
    add_attachment_to_asset,
    delete_attachment,
    get_attachment_abs_path,
)

from services.custom_fields_service import (
    list_field_defs,
    list_values_for_asset,
    bulk_set_values_for_asset,
)

# ✅ Asset service (RBAC+scope) + disposal workflow (best-effort import)
_ASSET_SVC_ERR = ""
try:
    from services.assets_service import (  # type: ignore
        get_asset_by_uid as svc_get_asset_by_uid,
        prepare_disposal as svc_prepare_disposal,
        approve_disposal as svc_approve_disposal,
        cancel_disposal as svc_cancel_disposal,
        dispose_from_case as svc_dispose_from_case,
    )
except Exception as e:  # pragma: no cover
    svc_get_asset_by_uid = None  # type: ignore
    svc_prepare_disposal = None  # type: ignore
    svc_approve_disposal = None  # type: ignore
    svc_cancel_disposal = None  # type: ignore
    svc_dispose_from_case = None  # type: ignore
    _ASSET_SVC_ERR = str(e)

# ✅ Metrologija: fail-safe import
_METRO_IMPORT_ERR = ""
try:
    from services.metrology_service import list_metrology_records_for_asset  # type: ignore
except Exception as e:  # pragma: no cover
    list_metrology_records_for_asset = None  # type: ignore
    _METRO_IMPORT_ERR = str(e)

# ✅ Kalendar: fail-safe import
_CAL_IMPORT_ERR = ""
try:
    from services.calendar_service import (  # type: ignore
        list_calendar_events_for_asset,
        add_calendar_event_for_asset,
        update_calendar_event,
        delete_calendar_event,
    )
except Exception as e:  # pragma: no cover
    list_calendar_events_for_asset = None  # type: ignore
    add_calendar_event_for_asset = None  # type: ignore
    update_calendar_event = None  # type: ignore
    delete_calendar_event = None  # type: ignore
    _CAL_IMPORT_ERR = str(e)

# ✅ Copy helpers
from ui.utils.table_copy import (
    wire_table_header_plus_copy,
    wire_table_selection_plus_copy,
    copy_selected_cells,
)

log = logging.getLogger(__name__)

_QS_ORG = "BazaS2"
_QS_APP = "BazaS2"
_QS_PREFIX = "ui/dialogs/AssetDetailDialog"
_OLD_GROUP = "ui.asset_detail_dialog"

_REQUIRED_FIELD_LABELS: List[Tuple[str, str]] = [
    ("name", "Naziv"),
    ("category", "Kategorija"),
    ("toc_number", "TOC"),
    ("nomenclature_no", "Nomenklaturni broj"),
    ("serial_number", "Serijski broj"),
    ("inventory_no", "Inventarski broj"),
    ("sector", "Sektor"),
    ("location", "Lokacija"),
    ("current_holder", "Zaduženo kod"),
]

_MISSING_COLOR = "#c0392b"
_HINT_COLOR = "#666"

# ✅ RBAC constants (fail-safe)
try:
    from core.rbac import (  # type: ignore
        PERM_METRO_MANAGE,
        PERM_METRO_VIEW,
        PERM_ASSETS_METRO_VIEW,
        PERM_ASSETS_EDIT,
        PERM_DISPOSAL_PREPARE,
        PERM_DISPOSAL_APPROVE,
        PERM_DISPOSAL_DISPOSE,
    )
except Exception:  # pragma: no cover
    PERM_METRO_MANAGE = "metrology.manage"
    PERM_METRO_VIEW = "metrology.view"
    PERM_ASSETS_METRO_VIEW = "assets.metrology.view"
    PERM_ASSETS_EDIT = "assets.edit"
    PERM_DISPOSAL_PREPARE = "disposal.prepare"
    PERM_DISPOSAL_APPROVE = "disposal.approve"
    PERM_DISPOSAL_DISPOSE = "disposal.dispose"

# ✅ Edit dialog (fail-safe)
_EDIT_IMPORT_ERR = ""
try:
    from ui.asset_edit_dialog import AssetEditDialog  # type: ignore
except Exception as e:  # pragma: no cover
    AssetEditDialog = None  # type: ignore
    _EDIT_IMPORT_ERR = str(e)


# -------------------- helpers --------------------

def _app_root() -> Path:
    return Path(__file__).resolve().parent.parent


def _resolve_db_path() -> Path:
    """
    Prefer core.db.get_db_path() (jedna istina), fallback na core.config.DB_FILE.
    """
    try:
        from core.db import get_db_path as _get_db_path  # type: ignore
        p = Path(str(_get_db_path() or "")).expanduser()
        if str(p):
            return p.resolve()
    except Exception:
        pass

    p2 = Path(DB_FILE)
    if not p2.is_absolute():
        p2 = (_app_root() / p2).resolve()
    return p2


def _norm(s: str) -> str:
    return (s or "").strip().casefold()


def _safe_int(v: Any, default: int = 0) -> int:
    try:
        return int(v)
    except Exception:
        return int(default)


def _qvariant_to_bytearray(v: Any) -> Optional[QByteArray]:
    if v is None:
        return None
    if isinstance(v, QByteArray):
        return v
    if isinstance(v, (bytes, bytearray)):
        return QByteArray(bytes(v))
    return None


def _set_clipboard_text(text: str) -> None:
    try:
        app = QApplication.instance()
        if app is None:
            return
        cb = app.clipboard()
        if cb is None:
            return
        cb.setText(str(text or ""))
    except Exception:
        pass


def _can_asset_edit() -> bool:
    try:
        if bool(can(PERM_ASSETS_EDIT)):
            return True
    except Exception:
        pass
    for p in ("assets.update", "assets.manage", "assets.write", "assets.edit", "assets.change_status"):
        try:
            if bool(can(p)):
                return True
        except Exception:
            continue
    return False


def _can_disposal_prepare() -> bool:
    try:
        return bool(can(PERM_DISPOSAL_PREPARE))
    except Exception:
        return False


def _can_disposal_approve() -> bool:
    try:
        return bool(can(PERM_DISPOSAL_APPROVE))
    except Exception:
        return False


def _can_disposal_dispose() -> bool:
    try:
        return bool(can(PERM_DISPOSAL_DISPOSE))
    except Exception:
        return False


def _current_row_any(table: QTableWidget) -> int:
    r = table.currentRow()
    if r >= 0:
        return r
    try:
        sm = table.selectionModel()
        if sm:
            idx = sm.selectedIndexes()
            if idx:
                return idx[0].row()
    except Exception:
        pass
    return -1


def _q_ident(name: str) -> str:
    n = str(name or "").replace('"', '""')
    return f'"{n}"'


def _safe_rel_under_root(rel_path: str) -> Optional[Path]:
    """
    Path traversal guard: relativna putanja mora ostati unutar app_root().
    """
    try:
        rel = Path(str(rel_path))
        if rel.is_absolute():
            return rel
        root = _app_root().resolve()
        resolved = (root / rel).resolve()
        try:
            resolved.relative_to(root)
        except Exception:
            return None
        return resolved
    except Exception:
        return None


def _read_asset_row(asset_uid: str) -> Dict[str, Any]:
    """
    SECURITY FIRST:
    - Primarno čitanje preko services.assets_service.get_asset_by_uid (RBAC + scope).
    - Fallback direktan DB read samo ako servis nije dostupan (kompatibilnost).
    """
    au = (asset_uid or "").strip()
    if not au:
        return {}

    # 1) Service path (RBAC + scope)
    if svc_get_asset_by_uid is not None:
        try:
            r = svc_get_asset_by_uid(asset_uid=au)  # type: ignore[misc]
            if isinstance(r, dict):
                out = dict(r)
                if "nomenclature_no" not in out and "nomenclature_number" in out:
                    out["nomenclature_no"] = out.get("nomenclature_number")
                return out
            return {}
        except PermissionError:
            raise
        except Exception as e:
            log.warning("assets_service.get_asset_by_uid failed (%s). Falling back to direct DB read.", e)

    # 2) Fallback direct DB read (best-effort)
    try:
        from core.db import connect_db as _connect_db  # type: ignore
    except Exception:
        _connect_db = None  # type: ignore

    def _fetch_from_conn(conn: sqlite3.Connection) -> Dict[str, Any]:
        t = conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='assets' LIMIT 1;").fetchone()
        if not t:
            return {}

        cols = [r[1] for r in conn.execute('PRAGMA table_info("assets");').fetchall()]
        if "asset_uid" not in cols:
            return {}

        def pick(*names: str) -> str:
            for n in names:
                if n in cols:
                    return n
            return ""

        col_name = pick("name", "asset_name")
        col_cat = pick("category", "cat")
        col_toc = pick("toc_number", "toc")
        col_sn = pick("serial_number", "serial")
        col_loc = pick("location", "loc")
        col_st = pick("status", "state")
        col_holder = pick("current_holder", "holder", "assigned_to")
        col_upd = pick("updated_at", "modified_at", "updated")
        col_cr = pick("created_at", "created", "inserted_at", "created_on")

        col_sector = pick("sector", "sektor", "org_unit", "unit", "department", "dept", "section")
        col_is_metro = pick("is_metrology", "is_metro", "metrology_flag", "metro_flag", "metrology_scope")

        col_model = pick("model", "device_model")
        col_vendor = pick("vendor", "manufacturer", "maker")
        col_inv = pick("inventory_no", "inv_no", "inventarski_broj")
        col_notes = pick("notes", "note", "napomena", "opis")

        col_nom = pick(
            "nomenclature_no", "nomenclature_number", "nomencl_no",
            "nomenklaturni_broj", "nomenkl_broj", "nomenklatura",
            "nom_no", "nom_number", "nomenclature",
        )

        sel_cols: List[str] = ["asset_uid"]
        for c in (
            col_toc, col_sn, col_nom,
            col_name, col_cat, col_st, col_holder, col_loc,
            col_cr, col_upd,
            col_sector, col_is_metro,
            col_model, col_vendor, col_inv, col_notes,
        ):
            if c:
                sel_cols.append(c)

        sel_sql = ", ".join(_q_ident(c) for c in sel_cols)
        row = conn.execute(f"SELECT {sel_sql} FROM assets WHERE asset_uid=? LIMIT 1;", (au,)).fetchone()
        if not row:
            return {}

        out: Dict[str, Any] = {sel_cols[i]: row[i] for i in range(len(sel_cols))}
        sector_val = out.get(col_sector, "") if col_sector else ""
        is_metro_val = out.get(col_is_metro, 0) if col_is_metro else 0
        is_metro_int = 1 if _safe_int(is_metro_val, 0) == 1 else 0

        return {
            "asset_uid": out.get("asset_uid", "") or "",
            "toc_number": out.get(col_toc, "") or "",
            "serial_number": out.get(col_sn, "") or "",
            "nomenclature_no": out.get(col_nom, "") or "",
            "name": out.get(col_name, "") or "",
            "category": out.get(col_cat, "") or "",
            "status": out.get(col_st, "") or "",
            "current_holder": out.get(col_holder, "") or "",
            "location": out.get(col_loc, "") or "",
            "created_at": out.get(col_cr, "") or "",
            "updated_at": out.get(col_upd, "") or "",
            "sector": str(sector_val or "").strip(),
            "is_metrology": is_metro_int,
            "model": out.get(col_model, "") or "",
            "vendor": out.get(col_vendor, "") or "",
            "inventory_no": out.get(col_inv, "") or "",
            "notes": out.get(col_notes, "") or "",
        }

    if _connect_db is not None:
        try:
            conn = _connect_db()
            try:
                return _fetch_from_conn(conn)
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass

    db_path = _resolve_db_path()
    if not db_path.exists():
        return {}

    conn = sqlite3.connect(db_path.as_posix())
    try:
        return _fetch_from_conn(conn)
    finally:
        try:
            conn.close()
        except Exception:
            pass


def _is_image_file(path: Path) -> bool:
    return (path.suffix or "").lower() in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp")


def _is_text_file(path: Path) -> bool:
    return (path.suffix or "").lower() in (".txt", ".log", ".csv", ".ini", ".json", ".xml")


def _is_docx(path: Path) -> bool:
    return (path.suffix or "").lower() == ".docx"


def _is_xlsx(path: Path) -> bool:
    return (path.suffix or "").lower() == ".xlsx"


def _is_pdf(path: Path) -> bool:
    return (path.suffix or "").lower() == ".pdf"


def _read_text_preview(path: Path, max_bytes: int = 200_000) -> str:
    try:
        tail = ""
        with path.open("rb") as f:
            data = f.read(max_bytes + 1)
        if len(data) > max_bytes:
            data = data[:max_bytes]
            tail = "\n\n--- PREVIEW SKRAĆEN (više sadržaja u fajlu) ---\n"
        try:
            txt = data.decode("utf-8", errors="replace")
        except Exception:
            txt = data.decode(errors="replace")
        return txt + tail
    except Exception as e:
        return f"[Ne mogu da pročitam TXT preview]\n{e}"


def _extract_docx_text(path: Path, max_chars: int = 200_000) -> str:
    try:
        import docx  # type: ignore
    except Exception as e:
        return f"[DOCX preview nije dostupan — nema python-docx]\n{e}"

    try:
        d = docx.Document(str(path))
        parts: List[str] = []
        total = 0
        for p in d.paragraphs:
            t = (p.text or "").strip()
            if t:
                parts.append(t)
                total += len(t)
            if total > max_chars:
                parts.append("\n--- PREVIEW SKRAĆEN ---")
                break
        if not parts:
            return "(DOCX) Nema tekstualnog sadržaja ili je prazan."
        return "\n".join(parts)[:max_chars]
    except Exception as e:
        return f"[Greška pri DOCX preview-u]\n{e}"


def _extract_xlsx_text(
    path: Path,
    max_chars: int = 200_000,
    max_sheets: int = 3,
    max_rows: int = 80,
    max_cols: int = 20,
) -> str:
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        return f"[XLSX preview nije dostupan — nema openpyxl]\n{e}"

    wb = None
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        out: List[str] = []
        total = 0
        for si, name in enumerate(wb.sheetnames[:max_sheets], start=1):
            ws = wb[name]
            hdr = f"=== Sheet {si}: {name} ==="
            out.append(hdr)
            total += len(hdr) + 1
            for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
                vals = []
                for v in row:
                    s = "" if v is None else str(v)
                    s = s.replace("\n", " ").strip()
                    vals.append(s)
                line = "\t".join(vals).rstrip()
                out.append(line)
                total += len(line) + 1
                if total > max_chars:
                    out.append("\n--- PREVIEW SKRAĆEN ---")
                    break
            out.append("")
            if total > max_chars:
                break
        txt = "\n".join(out).strip()
        return txt[:max_chars] if txt else "(XLSX) Nema čitljivog sadržaja u preview opsegu."
    except Exception as e:
        return f"[Greška pri XLSX preview-u]\n{e}"
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def _extract_pdf_text(path: Path, max_chars: int = 200_000, max_pages: int = 5) -> str:
    try:
        try:
            from pypdf import PdfReader  # type: ignore
        except Exception:
            from PyPDF2 import PdfReader  # type: ignore
    except Exception as e:
        return f"[PDF preview nije dostupan — nema pypdf/PyPDF2]\n{e}"

    try:
        r = PdfReader(str(path))
        out: List[str] = []
        total = 0
        for i, page in enumerate(r.pages[:max_pages], start=1):
            try:
                t = page.extract_text() or ""
            except Exception:
                t = ""
            t = t.strip()
            head = f"=== Strana {i} ==="
            out.append(head)
            total += len(head) + 1
            body = t if t else "(nema izvučenog teksta — možda je sken)"
            out.append(body)
            total += len(body) + 1
            out.append("")
            total += 1
            if total > max_chars:
                out.append("\n--- PREVIEW SKRAĆEN ---")
                break
        txt = "\n".join(out).strip()
        return txt[:max_chars] if txt else "(PDF) Nema čitljivog teksta (možda je sken)."
    except Exception as e:
        return f"[Greška pri PDF preview-u]\n{e}"


def _wire_table_full_copy(table: QTableWidget) -> None:
    try:
        table.setSelectionBehavior(QAbstractItemView.SelectItems)
        table.setSelectionMode(QAbstractItemView.ExtendedSelection)
    except Exception:
        pass
    try:
        wire_table_selection_plus_copy(table)
    except Exception:
        pass
    try:
        wire_table_header_plus_copy(table)
    except Exception:
        pass


def _load_timeline_rows(asset_uid: str, limit: int = 400) -> Tuple[List[Dict[str, Any]], str]:
    """
    Prefer core.db.list_asset_events_db (asset_events tabela).
    Fallback: legacy auto-detekcija tabela.
    """
    au = (asset_uid or "").strip()
    if not au:
        return [], "Nema asset_uid."

    # 1) Prefer core.db API
    try:
        from core.db import list_asset_events_db  # type: ignore
        rows = list_asset_events_db(au, limit=int(limit)) or []
        out: List[Dict[str, Any]] = []
        for rr in rows:
            out.append({
                "event_time": rr.get("event_time", ""),
                "actor": rr.get("actor", ""),
                "event_type": rr.get("event_type", ""),
                "details": rr.get("data_json", rr.get("data", "")),
                "source": rr.get("source", ""),
            })
        return out, f"asset_events: {len(out)} događaja"
    except Exception:
        pass

    # 2) Legacy fallback: direct SQL scan
    db_path = _resolve_db_path()
    if not db_path.exists():
        return [], "DB nije pronađena."

    try:
        conn = sqlite3.connect(db_path.as_posix())
    except Exception as e:
        return [], f"Ne mogu da otvorim DB: {e}"

    def table_exists(name: str) -> bool:
        try:
            r = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1;",
                (name,),
            ).fetchone()
            return bool(r)
        except Exception:
            return False

    def cols(name: str) -> List[str]:
        try:
            return [r[1] for r in conn.execute(f"PRAGMA table_info({_q_ident(name)});").fetchall()]
        except Exception:
            return []

    def pick(candidates: List[str], available: List[str]) -> str:
        for c in candidates:
            if c in available:
                return c
        return ""

    try:
        candidates = ["asset_events", "audit_log", "events", "timeline"]
        tname = next((t for t in candidates if table_exists(t)), "")
        if not tname:
            return [], "Timeline tabela nije pronađena."

        c = cols(tname)
        col_asset = pick(["asset_uid", "asset", "uid"], c)
        col_ts = pick(["created_at", "ts", "timestamp", "time", "event_time"], c)
        col_actor = pick(["actor", "user", "username", "performed_by", "by_user", "who"], c)
        col_action = pick(["action", "event", "event_type", "type", "op", "operation", "verb"], c)
        col_details = pick(["details", "detail", "message", "payload", "data_json", "data", "note", "changes", "description"], c)

        if not col_asset:
            return [], f"Timeline tabela '{tname}' nema asset_uid kolonu."

        sel: List[str] = [x for x in (col_ts, col_actor, col_action, col_details) if x]
        if not sel:
            sel = c[:]

        order = f"ORDER BY {_q_ident(col_ts)} DESC" if col_ts else "ORDER BY rowid DESC"
        sel_sql = ", ".join(_q_ident(x) for x in sel)
        q = f"SELECT {sel_sql} FROM {_q_ident(tname)} WHERE {_q_ident(col_asset)}=? {order} LIMIT ?;"
        rows = conn.execute(q, (au, int(limit))).fetchall()

        out: List[Dict[str, Any]] = []
        for r0 in rows:
            dct = {}
            for i, name in enumerate(sel):
                dct[name] = r0[i]
            out.append(dct)

        return out, f"{tname}: {len(out)} događaja"
    except Exception as e:
        return [], f"Timeline greška: {e}"
    finally:
        try:
            conn.close()
        except Exception:
            pass


# -------------------- preview worker --------------------

class _PreviewWorker(QObject):
    finished = Signal(int, str, str)  # job_id, path_str, text

    def __init__(self, job_id: int, path_str: str):
        super().__init__()
        self.job_id = int(job_id)
        self.path_str = str(path_str)

    def _interrupted(self) -> bool:
        try:
            th = QThread.currentThread()
            return bool(th.isInterruptionRequested())
        except Exception:
            return False

    def run(self) -> None:
        p = Path(self.path_str)
        try:
            if self._interrupted():
                txt = "Preview prekidan (promenjena selekcija)."
            elif (not p.exists()) or (not p.is_file()):
                txt = "Fajl ne postoji na disku (ili putanja nije dostupna)."
            elif _is_text_file(p):
                txt = _read_text_preview(p)
            elif _is_docx(p):
                txt = _extract_docx_text(p)
            elif _is_xlsx(p):
                txt = _extract_xlsx_text(p)
            elif _is_pdf(p):
                txt = _extract_pdf_text(p)
            else:
                txt = f"Preview nije podržan za tip fajla: {p.suffix}\n\nKoristi 'Otvori'."
        except Exception as e:
            txt = f"[Greška pri generisanju preview-a]\n{e}"

        try:
            self.finished.emit(self.job_id, self.path_str, txt)
        except Exception:
            pass

# (FILENAME: ui/asset_detail_dialog.py - END PART 1/4)


# FILENAME: ui/asset_detail_dialog.py
# (FILENAME: ui/asset_detail_dialog.py - START PART 2/4)

class AssetDetailDialog(QDialog):
    def __init__(self, asset_uid: str, parent=None):
        super().__init__(parent)

        self.setObjectName("AssetDetailDialog")
        try:
            self.setAttribute(Qt.WA_DeleteOnClose, True)
        except Exception:
            pass

        self.asset_uid = (asset_uid or "").strip()
        self.setWindowTitle(f"Detalji sredstva — {self.asset_uid}")

        try:
            self._qs: Optional[QSettings] = QSettings(_QS_ORG, _QS_APP)
        except Exception:
            self._qs = None

        self._restored_once = False
        self._base_styles: Dict[int, str] = {}
        self._closing = False

        self._asset_row: Dict[str, Any] = {}
        self._cal_selected_id: int = 0

        self._cal_split: Optional[QSplitter] = None
        self._att_split: Optional[QSplitter] = None

        self._att_rows: List[Dict[str, Any]] = []
        self._att_selected: Optional[Dict[str, Any]] = None
        self._preview_img_path: Optional[Path] = None

        self._last_find_text: str = ""
        self._last_find_cursor: Optional[QTextCursor] = None

        self._preview_job_id: int = 0
        self._preview_job_path: str = ""
        self._preview_thread: Optional[QThread] = None
        self._preview_worker: Optional[_PreviewWorker] = None

        self._apply_window_chrome()
        self.resize(1020, 720)
        self.setMinimumSize(640, 420)

        # -------------------- HEADER --------------------
        header = QGroupBox("Sredstvo")
        hg = QGridLayout(header)
        hg.setContentsMargins(10, 10, 10, 10)

        self.lb_uid = QLabel(self.asset_uid)
        self.lb_uid.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.lb_uid.setStyleSheet("font-weight: 700;")

        self.lb_name = QLabel("")
        self.lb_name.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.lb_name.setStyleSheet("font-weight: 600;")

        self.lb_status = QLabel("")
        self.lb_status.setTextInteractionFlags(Qt.TextSelectableByMouse)

        self.lb_disposal = QLabel("")  # disposal badge
        self.lb_disposal.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.lb_disposal.setStyleSheet("color:#999; font-weight:600;")

        self.lb_cat = QLabel("")
        self.lb_cat.setTextInteractionFlags(Qt.TextSelectableByMouse)

        self.lb_nom = QLabel("")
        self.lb_nom.setTextInteractionFlags(Qt.TextSelectableByMouse)

        self.btn_scrap = QPushButton("Rashod")
        self.btn_scrap.setToolTip("Rashod workflow (Priprema/Odobri/Rashoduj) + legacy fallback.")
        self.btn_scrap.clicked.connect(self._open_disposal_menu)
        self.btn_scrap.setEnabled(_can_asset_edit() or _can_disposal_prepare() or _can_disposal_dispose())
        self.btn_scrap.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        self.btn_edit = QPushButton("Izmeni")
        self.btn_edit.setToolTip("Izmeni detalje sredstva.")
        self.btn_edit.clicked.connect(self._edit_asset)
        self.btn_edit.setEnabled(_can_asset_edit() and (AssetEditDialog is not None))
        self.btn_edit.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        self.lb_toc = QLabel("")
        self.lb_sn = QLabel("")
        self.lb_holder = QLabel("")
        self.lb_loc = QLabel("")
        self.lb_created = QLabel("")
        self.lb_updated = QLabel("")
        self.lb_sector = QLabel("")
        self.lb_is_metro = QLabel("")
        self.lb_inventory = QLabel("")

        self.lb_missing = QLabel("")
        self.lb_missing.setWordWrap(True)
        self.lb_missing.setStyleSheet(f"color:{_MISSING_COLOR}; font-weight:600;")

        for w in (
            self.lb_name, self.lb_status, self.lb_disposal, self.lb_cat, self.lb_nom,
            self.lb_toc, self.lb_sn, self.lb_holder, self.lb_loc,
            self.lb_created, self.lb_updated, self.lb_sector,
            self.lb_is_metro, self.lb_inventory
        ):
            self._remember_style(w)

        def _k(label: str) -> QLabel:
            l = QLabel(label)
            l.setStyleSheet(f"color:{_HINT_COLOR};")
            return l

        r = 0
        hg.addWidget(_k("UID"), r, 0); hg.addWidget(self.lb_uid, r, 1)
        hg.addWidget(_k("Status"), r, 2); hg.addWidget(self.lb_status, r, 3)
        hg.addWidget(self.btn_scrap, r, 4)
        hg.addWidget(self.btn_edit, r, 5)
        r += 1

        hg.addWidget(_k("Rashod"), r, 0); hg.addWidget(self.lb_disposal, r, 1, 1, 5)
        r += 1

        hg.addWidget(_k("Naziv"), r, 0); hg.addWidget(self.lb_name, r, 1, 1, 5)
        r += 1

        hg.addWidget(_k("Kategorija"), r, 0); hg.addWidget(self.lb_cat, r, 1)
        hg.addWidget(_k("TOC"), r, 2); hg.addWidget(self.lb_toc, r, 3)
        hg.addWidget(_k("Serijski"), r, 4); hg.addWidget(self.lb_sn, r, 5)
        r += 1

        hg.addWidget(_k("Nomenkl. broj"), r, 0); hg.addWidget(self.lb_nom, r, 1)
        hg.addWidget(_k("Inventarski broj"), r, 2); hg.addWidget(self.lb_inventory, r, 3)
        r += 1

        hg.addWidget(_k("Zaduženo kod"), r, 0); hg.addWidget(self.lb_holder, r, 1)
        hg.addWidget(_k("Lokacija"), r, 2); hg.addWidget(self.lb_loc, r, 3)
        hg.addWidget(_k("Uneto"), r, 4); hg.addWidget(self.lb_created, r, 5)
        r += 1

        hg.addWidget(_k("Ažurirano"), r, 0); hg.addWidget(self.lb_updated, r, 1)
        hg.addWidget(_k("Sektor"), r, 2); hg.addWidget(self.lb_sector, r, 3)
        hg.addWidget(_k("Metrologija flag"), r, 4); hg.addWidget(self.lb_is_metro, r, 5)
        r += 1

        hg.addWidget(self.lb_missing, r, 0, 1, 6)
        hg.setColumnStretch(1, 2)
        hg.setColumnStretch(3, 2)
        hg.setColumnStretch(5, 1)

        # -------------------- TABOVI --------------------
        self.tabs = QTabWidget()

        self.details_tab = QWidget()
        self.timeline_tab = QWidget()
        self.calendar_tab = QWidget()
        self.met_tab = QWidget()
        self.custom_tab = QWidget()
        self.attach_tab = QWidget()

        self._build_details_tab()
        self._build_timeline_tab()
        self._build_calendar_tab()
        self._build_metrology_tab()
        self._build_custom_fields_tab()
        self._build_attachments_tab()  # ✅ sada je kompletno i stabilno

        self.tabs.addTab(self.details_tab, "Detalji")
        self.tabs.addTab(self.timeline_tab, "Timeline")
        self.tabs.addTab(self.calendar_tab, "Kalendar")
        self.tabs.addTab(self.met_tab, "Metrologija")
        self.tabs.addTab(self.custom_tab, "Dodatna polja")
        self.tabs.addTab(self.attach_tab, "Prilozi")

        self.btns = QDialogButtonBox(QDialogButtonBox.Close)
        close_btn = self.btns.button(QDialogButtonBox.Close)
        if close_btn is not None:
            close_btn.setText("Zatvori")
        self.btns.rejected.connect(self.reject)
        self.btns.accepted.connect(self.reject)

        root = QVBoxLayout(self)
        root.addWidget(header)
        root.addWidget(self.tabs, 1)
        root.addWidget(self.btns)

        # SECURITY: ako nema pristup asset-u, zatvori dijalog fail-closed
        try:
            self._reload_all()
        except PermissionError as e:
            QMessageBox.warning(self, "RBAC / Scope", f"Nemaš pravo da vidiš ovo sredstvo.\n\n{e}")
            QTimer.singleShot(0, self.reject)

    # -------------------- window chrome / persist --------------------

    def _settings_key(self, suffix: str) -> str:
        return f"{_QS_PREFIX}/{suffix}"

    def _apply_window_chrome(self) -> None:
        try:
            self.setSizeGripEnabled(True)
        except Exception:
            pass
        try:
            self.setWindowFlag(Qt.WindowMinimizeButtonHint, True)
            self.setWindowFlag(Qt.WindowMaximizeButtonHint, True)
            self.setWindowFlag(Qt.WindowContextHelpButtonHint, False)
            self.setWindowFlag(Qt.MSWindowsFixedSizeDialogHint, False)
        except Exception:
            pass
        try:
            self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.setMaximumSize(16777215, 16777215)
        except Exception:
            pass

    def _restore_ui_state(self) -> None:
        if not self._qs:
            return

        restored_geo = False

        try:
            geo_norm = _qvariant_to_bytearray(self._qs.value(self._settings_key("normal_geometry"), None))
            geo = _qvariant_to_bytearray(self._qs.value(self._settings_key("geometry"), None))
            use_geo = geo_norm if (geo_norm and len(geo_norm) > 12) else geo
            if use_geo and len(use_geo) > 12:
                self.restoreGeometry(use_geo)
                restored_geo = True
        except Exception:
            pass

        try:
            idx = self._qs.value(self._settings_key("tab_index"), None)
            if idx is not None:
                self.tabs.setCurrentIndex(_safe_int(idx, 0))
        except Exception:
            pass

        try:
            was_max = self._qs.value(self._settings_key("maximized"), False)
            if str(was_max).lower() in ("1", "true", "yes", "da"):
                self.showMaximized()
        except Exception:
            pass

        try:
            if self._cal_split is not None:
                st = _qvariant_to_bytearray(self._qs.value(self._settings_key("split_calendar"), None))
                if st and len(st) > 4:
                    self._cal_split.restoreState(st)
        except Exception:
            pass

        try:
            if self._att_split is not None:
                st = _qvariant_to_bytearray(self._qs.value(self._settings_key("split_attachments"), None))
                if st and len(st) > 4:
                    self._att_split.restoreState(st)
        except Exception:
            pass

        # backward compat — apply only if new not restored
        try:
            self._qs.beginGroup(_OLD_GROUP)

            if not restored_geo:
                geo_old = self._qs.value("geometry", None)
                if geo_old is not None:
                    try:
                        self.restoreGeometry(geo_old)  # type: ignore[arg-type]
                    except Exception:
                        pass

            was_max_old = self._qs.value("maximized", False)
            if str(was_max_old).lower() in ("1", "true", "yes", "da"):
                try:
                    self.showMaximized()
                except Exception:
                    pass

            try:
                if self._cal_split is not None:
                    sizes = self._qs.value("cal_split_sizes", None)
                    if sizes and isinstance(sizes, (list, tuple)) and len(sizes) >= 2:
                        self._cal_split.setSizes([int(x) for x in sizes])
            except Exception:
                pass

            try:
                if self._att_split is not None:
                    sizes = self._qs.value("att_split_sizes", None)
                    if sizes and isinstance(sizes, (list, tuple)) and len(sizes) >= 2:
                        self._att_split.setSizes([int(x) for x in sizes])
            except Exception:
                pass
        finally:
            try:
                self._qs.endGroup()
            except Exception:
                pass

    def _save_ui_state(self) -> None:
        if not self._qs:
            return

        try:
            self._qs.setValue(self._settings_key("geometry"), self.saveGeometry())
        except Exception:
            pass

        try:
            if self.isMaximized():
                try:
                    self.showNormal()
                    self._qs.setValue(self._settings_key("normal_geometry"), self.saveGeometry())
                    self.showMaximized()
                except Exception:
                    pass
            else:
                self._qs.setValue(self._settings_key("normal_geometry"), self.saveGeometry())
        except Exception:
            pass

        try:
            self._qs.setValue(self._settings_key("maximized"), bool(self.isMaximized()))
        except Exception:
            pass

        try:
            self._qs.setValue(self._settings_key("tab_index"), int(self.tabs.currentIndex()))
        except Exception:
            pass

        try:
            if self._cal_split is not None:
                self._qs.setValue(self._settings_key("split_calendar"), self._cal_split.saveState())
        except Exception:
            pass

        try:
            if self._att_split is not None:
                self._qs.setValue(self._settings_key("split_attachments"), self._att_split.saveState())
        except Exception:
            pass

        # optional old format
        try:
            self._qs.beginGroup(_OLD_GROUP)
            self._qs.setValue("geometry", self.saveGeometry())
            self._qs.setValue("maximized", bool(self.isMaximized()))
            if self._cal_split is not None:
                self._qs.setValue("cal_split_sizes", self._cal_split.sizes())
            if self._att_split is not None:
                self._qs.setValue("att_split_sizes", self._att_split.sizes())
        finally:
            try:
                self._qs.endGroup()
            except Exception:
                pass

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        if self._restored_once:
            return
        self._restored_once = True
        QTimer.singleShot(0, self._restore_ui_state)

    def closeEvent(self, event) -> None:  # type: ignore[override]
        self._closing = True
        try:
            self._stop_preview_thread()
        except Exception:
            pass
        try:
            self._save_ui_state()
        except Exception:
            pass
        try:
            super().closeEvent(event)
        except Exception:
            pass

    def resizeEvent(self, event) -> None:  # type: ignore[override]
        super().resizeEvent(event)
        try:
            if self._preview_img_path and getattr(self, "preview_stack", None) is not None and self.preview_stack.currentIndex() == 1:
                self._render_image_preview(self._preview_img_path)
        except Exception:
            pass

    # -------------------- style helpers --------------------

    def _remember_style(self, w: QWidget) -> None:
        try:
            self._base_styles[id(w)] = w.styleSheet() or ""
        except Exception:
            pass

    def _set_missing_style(self, w: QWidget, missing: bool) -> None:
        base = self._base_styles.get(id(w), "")
        if not missing:
            try:
                w.setStyleSheet(base)
            except Exception:
                pass
            return
        try:
            w.setStyleSheet(base + f"; color:{_MISSING_COLOR}; font-weight:700;")
        except Exception:
            pass

    def _set_text_or_dash(self, w: QLabel, value: Any) -> None:
        s = str(value or "").strip()
        try:
            w.setText(s if s else "—")
        except Exception:
            pass

    # -------------------- tab builders --------------------

    def _build_details_tab(self) -> None:
        lay = QVBoxLayout(self.details_tab)

        box = QGroupBox("Detaljne informacije")
        g = QGridLayout(box)
        g.setContentsMargins(10, 10, 10, 10)

        self.det_uid = QLabel("")
        self.det_toc = QLabel("")
        self.det_sn = QLabel("")
        self.det_nom = QLabel("")
        self.det_name = QLabel("")
        self.det_cat = QLabel("")
        self.det_status = QLabel("")
        self.det_holder = QLabel("")
        self.det_loc = QLabel("")
        self.det_sector = QLabel("")
        self.det_is_metro = QLabel("")
        self.det_created = QLabel("")
        self.det_updated = QLabel("")
        self.det_vendor = QLabel("")
        self.det_model = QLabel("")
        self.det_inventory = QLabel("")

        self.det_notes = QPlainTextEdit()
        self.det_notes.setReadOnly(True)
        self.det_notes.setMinimumHeight(80)
        self.det_notes.setMaximumHeight(260)

        self.det_missing = QLabel("")
        self.det_missing.setWordWrap(True)
        self.det_missing.setStyleSheet(f"color:{_MISSING_COLOR}; font-weight:600;")

        for w in (
            self.det_uid, self.det_toc, self.det_sn, self.det_nom, self.det_name, self.det_cat,
            self.det_status, self.det_holder, self.det_loc, self.det_sector, self.det_is_metro,
            self.det_created, self.det_updated, self.det_vendor, self.det_model, self.det_inventory,
        ):
            try:
                w.setTextInteractionFlags(Qt.TextSelectableByMouse)
            except Exception:
                pass
            self._remember_style(w)

        def k(text: str) -> QLabel:
            l = QLabel(text)
            l.setStyleSheet(f"color:{_HINT_COLOR};")
            return l

        r = 0
        g.addWidget(k("UID"), r, 0); g.addWidget(self.det_uid, r, 1)
        g.addWidget(k("Status"), r, 2); g.addWidget(self.det_status, r, 3)
        g.addWidget(k("Uneto"), r, 4); g.addWidget(self.det_created, r, 5)
        r += 1

        g.addWidget(k("Ažurirano"), r, 0); g.addWidget(self.det_updated, r, 1)
        g.addWidget(k("Sektor"), r, 2); g.addWidget(self.det_sector, r, 3)
        g.addWidget(k("Metrologija flag"), r, 4); g.addWidget(self.det_is_metro, r, 5)
        r += 1

        g.addWidget(k("Naziv"), r, 0); g.addWidget(self.det_name, r, 1, 1, 5)
        r += 1

        g.addWidget(k("Kategorija"), r, 0); g.addWidget(self.det_cat, r, 1)
        g.addWidget(k("TOC"), r, 2); g.addWidget(self.det_toc, r, 3)
        g.addWidget(k("Serijski"), r, 4); g.addWidget(self.det_sn, r, 5)
        r += 1

        g.addWidget(k("Nomenkl. broj"), r, 0); g.addWidget(self.det_nom, r, 1)
        g.addWidget(k("Inventarski broj"), r, 2); g.addWidget(self.det_inventory, r, 3, 1, 3)
        r += 1

        g.addWidget(k("Zaduženo kod"), r, 0); g.addWidget(self.det_holder, r, 1)
        g.addWidget(k("Lokacija"), r, 2); g.addWidget(self.det_loc, r, 3)
        g.addWidget(k("Proizvođač"), r, 4); g.addWidget(self.det_vendor, r, 5)
        r += 1

        g.addWidget(k("Model"), r, 0); g.addWidget(self.det_model, r, 1)
        r += 1

        g.addWidget(self.det_missing, r, 0, 1, 6)
        r += 1

        g.addWidget(k("Napomena/Opis"), r, 0, Qt.AlignTop)
        g.addWidget(self.det_notes, r, 1, 1, 5)

        g.setColumnStretch(1, 2)
        g.setColumnStretch(3, 2)
        g.setColumnStretch(5, 1)

        lay.addWidget(box)
        lay.addStretch(1)

# (FILENAME: ui/asset_detail_dialog.py - END PART 2/4)

# FILENAME: ui/asset_detail_dialog.py
# (FILENAME: ui/asset_detail_dialog.py - START PART 3/4)

    def _build_timeline_tab(self) -> None:
        lay = QVBoxLayout(self.timeline_tab)

        top = QHBoxLayout()
        self.lb_tl_info = QLabel("")
        self.lb_tl_info.setStyleSheet(f"color:{_HINT_COLOR};")
        self.btn_tl_refresh = QPushButton("Osveži")
        self.btn_tl_refresh.clicked.connect(self._load_timeline)

        top.addWidget(self.btn_tl_refresh)
        top.addStretch(1)
        top.addWidget(self.lb_tl_info)

        self.tbl_tl = QTableWidget(0, 4)
        self.tbl_tl.setHorizontalHeaderLabels(["Vreme", "Korisnik", "Akcija", "Detalji"])
        self.tbl_tl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_tl.setAlternatingRowColors(True)
        self.tbl_tl.horizontalHeader().setStretchLastSection(True)
        _wire_table_full_copy(self.tbl_tl)

        lay.addLayout(top)
        lay.addWidget(self.tbl_tl, 1)

    def _build_calendar_tab(self) -> None:
        lay = QVBoxLayout(self.calendar_tab)

        top = QHBoxLayout()
        self.lb_cal_info = QLabel("")
        self.lb_cal_info.setStyleSheet(f"color:{_HINT_COLOR};")

        self.btn_cal_refresh = QPushButton("Osveži")
        self.btn_cal_add = QPushButton("Dodaj")
        self.btn_cal_edit = QPushButton("Izmeni")
        self.btn_cal_del = QPushButton("Obriši")

        self.btn_cal_edit.setEnabled(False)
        self.btn_cal_del.setEnabled(False)

        top.addWidget(self.btn_cal_refresh)
        top.addWidget(self.btn_cal_add)
        top.addWidget(self.btn_cal_edit)
        top.addWidget(self.btn_cal_del)
        top.addStretch(1)
        top.addWidget(self.lb_cal_info)

        self.cal = QCalendarWidget()
        self.cal.setGridVisible(True)
        self.cal.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.tbl_cal = QTableWidget(0, 4)
        self.tbl_cal.setHorizontalHeaderLabels(["ID", "Vreme", "Tip", "Opis"])
        self.tbl_cal.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_cal.setAlternatingRowColors(True)
        self.tbl_cal.horizontalHeader().setStretchLastSection(True)
        _wire_table_full_copy(self.tbl_cal)

        split = QSplitter(Qt.Horizontal)
        split.setChildrenCollapsible(True)
        split.addWidget(self.cal)
        split.addWidget(self.tbl_cal)
        split.setStretchFactor(0, 1)
        split.setStretchFactor(1, 2)
        self._cal_split = split

        lay.addLayout(top)
        lay.addWidget(split, 1)

        self.btn_cal_refresh.clicked.connect(self._load_calendar_for_selected_date)
        self.cal.selectionChanged.connect(self._load_calendar_for_selected_date)
        self.tbl_cal.itemSelectionChanged.connect(self._sync_calendar_buttons)
        self.tbl_cal.cellDoubleClicked.connect(lambda r, c: self._calendar_edit())
        self.btn_cal_add.clicked.connect(self._calendar_add)
        self.btn_cal_edit.clicked.connect(self._calendar_edit)
        self.btn_cal_del.clicked.connect(self._calendar_delete)

    def _build_metrology_tab(self) -> None:
        lay = QVBoxLayout(self.met_tab)

        self.met_box = QGroupBox("Metrologija (etaloniranja/kalibracije za ovo sredstvo)")
        met_lay = QVBoxLayout(self.met_box)

        met_top = QHBoxLayout()
        self.lb_met_info = QLabel("")
        self.lb_met_info.setStyleSheet(f"color:{_HINT_COLOR};")
        self.cb_met_warn = QComboBox()
        self.cb_met_warn.addItems(["7", "14", "30", "60", "90"])
        self.cb_met_warn.setCurrentText("30")
        self.btn_met_refresh = QPushButton("Osveži")
        self.btn_met_open = QPushButton("Detalji zapisa")
        self.btn_met_open.setEnabled(False)

        met_top.addWidget(QLabel("Alarm prag (dana):"))
        met_top.addWidget(self.cb_met_warn)
        met_top.addWidget(self.btn_met_refresh)
        met_top.addStretch(1)
        met_top.addWidget(self.lb_met_info)
        met_top.addWidget(self.btn_met_open)

        self.tbl_met = QTableWidget(0, 7)
        self.tbl_met.setHorizontalHeaderLabels(["Status", "Met UID", "Tip", "Datum", "Važi do", "Izvršilac/Lab", "Sertifikat"])
        self.tbl_met.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_met.setAlternatingRowColors(True)
        self.tbl_met.horizontalHeader().setStretchLastSection(True)
        _wire_table_full_copy(self.tbl_met)

        self.tbl_met.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tbl_met.customContextMenuRequested.connect(self._on_met_context_menu)

        met_lay.addLayout(met_top)
        met_lay.addWidget(self.tbl_met, 1)

        lay.addWidget(self.met_box, 1)

        self.btn_met_refresh.clicked.connect(self._load_metrology)
        self.cb_met_warn.currentIndexChanged.connect(self._load_metrology)
        self.tbl_met.itemSelectionChanged.connect(self._sync_met_buttons)
        self.tbl_met.cellDoubleClicked.connect(lambda r, c: self._open_met_details())
        self.btn_met_open.clicked.connect(self._open_met_details)

    def _build_custom_fields_tab(self) -> None:
        lay = QVBoxLayout(self.custom_tab)

        self.custom_box = QGroupBox("Dodatna polja (Admin definiše u meniju: Prilagođena polja)")
        self.custom_form = QFormLayout(self.custom_box)
        self.custom_form.setContentsMargins(10, 8, 10, 10)

        self._custom_defs: List[Dict[str, Any]] = []
        self._custom_widgets: Dict[str, Any] = {}

        self.btn_custom_save = QPushButton("Snimi dodatna polja")
        self.btn_custom_save.clicked.connect(self._save_custom_fields)
        self.custom_form.addRow("", self.btn_custom_save)

        lay.addWidget(self.custom_box, 1)
        lay.addStretch(1)

    # ✅ Prilozi tab (UI builder) — metode za akcije/preview su u Part 4
    def _build_attachments_tab(self) -> None:
        lay = QVBoxLayout(self.attach_tab)

        att_box = QGroupBox("Prilozi")
        att_lay = QVBoxLayout(att_box)

        top_btns = QHBoxLayout()
        self.btn_add = QPushButton("Dodaj prilog")
        self.btn_open = QPushButton("Otvori")
        self.btn_open_folder = QPushButton("Otvori folder")
        self.btn_del = QPushButton("Obriši")

        self.btn_add.setToolTip("Dodaj fajl kao prilog sredstvu.")
        self.btn_open.setToolTip("Otvori izabrani prilog u podrazumevanoj aplikaciji.")
        self.btn_open_folder.setToolTip("Otvori folder u kojem se nalazi izabrani prilog.")
        self.btn_del.setToolTip("Obriši izabrani prilog (zapis + fajl po pravilima servisa).")

        self.btn_open.setEnabled(False)
        self.btn_open_folder.setEnabled(False)
        self.btn_del.setEnabled(False)

        top_btns.addWidget(self.btn_add)
        top_btns.addWidget(self.btn_open)
        top_btns.addWidget(self.btn_open_folder)
        top_btns.addWidget(self.btn_del)
        top_btns.addStretch(1)

        # path bar
        path_bar = QHBoxLayout()
        path_bar.setContentsMargins(0, 0, 0, 0)

        self.ed_att_path = QLineEdit()
        self.ed_att_path.setReadOnly(True)
        self.ed_att_path.setPlaceholderText("Putanja izabranog priloga…")

        self.btn_copy_path = QPushButton("Kopiraj putanju")
        self.btn_copy_path.setEnabled(False)
        self.btn_copy_path.setToolTip("Kopira putanju izabranog priloga u clipboard.")
        self.btn_copy_path.clicked.connect(lambda: _set_clipboard_text(self.ed_att_path.text() or ""))

        path_bar.addWidget(QLabel("Putanja:"))
        path_bar.addWidget(self.ed_att_path, 1)
        path_bar.addWidget(self.btn_copy_path)

        # table
        self.tbl = QTableWidget(0, 4)
        self.tbl.setHorizontalHeaderLabels(["R.br.", "Naziv", "Kreirano", "Napomena"])
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self.tbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        _wire_table_full_copy(self.tbl)
        try:
            self.tbl.setSortingEnabled(True)
        except Exception:
            pass

        self.tbl.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tbl.customContextMenuRequested.connect(self._on_tbl_context_menu)  # ✅ FIX: metoda postoji u Part 4

        # preview + find
        self.preview_stack = QTabWidget()
        self.preview_stack.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.ed_find = QLineEdit()
        self.ed_find.setPlaceholderText("Pretraga u preview-u (Enter = Sledeće)")
        self.btn_find = QPushButton("Nađi")
        self.btn_next = QPushButton("Sledeće")
        self.btn_prev = QPushButton("Prethodno")
        self.lb_find = QLabel("")

        self.btn_find.setEnabled(False)
        self.btn_next.setEnabled(False)
        self.btn_prev.setEnabled(False)

        find_bar = QHBoxLayout()
        find_bar.setContentsMargins(0, 0, 0, 0)
        find_bar.addWidget(QLabel("🔎"))
        find_bar.addWidget(self.ed_find, 1)
        find_bar.addWidget(self.btn_find)
        find_bar.addWidget(self.btn_prev)
        find_bar.addWidget(self.btn_next)
        find_bar.addWidget(self.lb_find)

        find_wrap = QWidget()
        find_wrap.setLayout(find_bar)
        find_wrap.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.preview_info = QLabel(
            "Izaberi prilog iz liste.\n\n"
            "• Slike: prikaz.\n"
            "• TXT/DOCX/XLSX/PDF: tekstualni preview + pretraga (best-effort).\n\n"
            "Tip: Ctrl+F fokusira polje pretrage (dok si na tabu Prilozi)."
        )
        self.preview_info.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self.preview_info.setWordWrap(True)
        self.preview_info.setStyleSheet("padding: 8px;")

        w_info = QWidget()
        li = QVBoxLayout(w_info)
        li.setContentsMargins(0, 0, 0, 0)
        li.addWidget(self.preview_info, 1)

        self.preview_img = QLabel()
        self.preview_img.setAlignment(Qt.AlignCenter)
        self.preview_img.setText("Nema pregleda.")
        self.preview_img.setStyleSheet("background:#111; color:#ddd; border:1px solid #333;")
        self.preview_img.setMinimumHeight(160)
        self.preview_img.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        w_img = QWidget()
        limg = QVBoxLayout(w_img)
        limg.setContentsMargins(0, 0, 0, 0)
        limg.addWidget(self.preview_img, 1)

        self.preview_txt = QPlainTextEdit()
        self.preview_txt.setReadOnly(True)
        self.preview_txt.setPlaceholderText("Tekstualni preview...")
        self.preview_txt.setStyleSheet("font-family: Consolas, 'Courier New', monospace;")
        try:
            self.preview_txt.setLineWrapMode(QPlainTextEdit.NoWrap)  # type: ignore[attr-defined]
        except Exception:
            pass

        w_txt = QWidget()
        ltxt = QVBoxLayout(w_txt)
        ltxt.setContentsMargins(0, 0, 0, 0)
        ltxt.addWidget(self.preview_txt, 1)

        self.preview_stack.addTab(w_info, "Info")
        self.preview_stack.addTab(w_img, "Slika")
        self.preview_stack.addTab(w_txt, "Tekst")
        self.preview_stack.setCurrentIndex(0)
        self.preview_stack.currentChanged.connect(lambda _: self._sync_find_enabled())

        preview_panel = QWidget()
        preview_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        preview_lay = QVBoxLayout(preview_panel)
        preview_lay.setContentsMargins(0, 0, 0, 0)
        preview_lay.addWidget(find_wrap)
        preview_lay.addWidget(self.preview_stack, 1)

        self.ed_note = QPlainTextEdit()
        self.ed_note.setPlaceholderText("Napomena za prilog (opciono) — upisuje se u bazu uz prilog.")
        self.ed_note.setMinimumHeight(48)
        self.ed_note.setMaximumHeight(140)

        split = QSplitter(Qt.Horizontal)
        split.setChildrenCollapsible(True)
        split.addWidget(self.tbl)
        split.addWidget(preview_panel)
        split.setStretchFactor(0, 3)
        split.setStretchFactor(1, 2)
        self._att_split = split

        att_lay.addLayout(top_btns)
        att_lay.addLayout(path_bar)
        att_lay.addWidget(split, 1)
        att_lay.addWidget(self.ed_note)

        lay.addWidget(att_box, 1)

        # signals
        self.btn_add.clicked.connect(self.add_attachment)
        self.btn_open.clicked.connect(self.open_attachment)
        self.btn_open_folder.clicked.connect(self.open_attachment_folder)
        self.btn_del.clicked.connect(self.delete_attachment_ui)

        self.tbl.itemSelectionChanged.connect(self._on_att_selection_changed)
        self.tbl.cellDoubleClicked.connect(lambda r, c: self.open_attachment())

        self.btn_find.clicked.connect(self._find_first)
        self.btn_next.clicked.connect(self._find_next)
        self.btn_prev.clicked.connect(self._find_prev)
        self.ed_find.returnPressed.connect(self._find_next)
        self.preview_txt.textChanged.connect(self._sync_find_enabled)
        self.ed_find.textChanged.connect(self._sync_find_enabled)

        # debounce selection -> preview
        try:
            self._att_preview_timer = QTimer(self)
            self._att_preview_timer.setSingleShot(True)
            self._att_preview_timer.timeout.connect(self._render_selected_attachment_preview)
        except Exception:
            self._att_preview_timer = None  # type: ignore

    # -------------------- reload --------------------

    def _reload_all(self) -> None:
        self._load_asset()  # može baciti PermissionError (fail-closed)
        self._refresh_disposal_badge()
        self._load_timeline()
        self._apply_calendar_scope_ui()
        self._load_calendar_for_selected_date()
        self._apply_metrology_scope_ui()
        self._load_metrology()
        self._load_custom_fields()
        self._load_attachments(show_errors=False)

    # -------------------- missing fields --------------------

    def _missing_fields(self) -> List[str]:
        missing: List[str] = []
        for key, label in _REQUIRED_FIELD_LABELS:
            val = self._asset_row.get(key, "")
            if not str(val or "").strip():
                missing.append(label)
        return missing

    def _apply_missing_warnings(self) -> None:
        missing = self._missing_fields()
        if missing:
            msg = "⚠ Nedostaju: " + ", ".join(missing)
            self.lb_missing.setText(msg)
            self.lb_missing.setVisible(True)
            self.det_missing.setText(msg)
            self.det_missing.setVisible(True)
        else:
            self.lb_missing.setText("")
            self.lb_missing.setVisible(False)
            self.det_missing.setText("")
            self.det_missing.setVisible(False)

        mapping: Dict[str, List[QLabel]] = {
            "name": [self.lb_name, self.det_name],
            "category": [self.lb_cat, self.det_cat],
            "toc_number": [self.lb_toc, self.det_toc],
            "nomenclature_no": [self.lb_nom, self.det_nom],
            "serial_number": [self.lb_sn, self.det_sn],
            "inventory_no": [self.lb_inventory, self.det_inventory],
            "sector": [self.lb_sector, self.det_sector],
            "location": [self.lb_loc, self.det_loc],
            "current_holder": [self.lb_holder, self.det_holder],
        }
        for key, widgets in mapping.items():
            val = self._asset_row.get(key, "")
            is_missing = not bool(str(val or "").strip())
            for w in widgets:
                self._set_missing_style(w, is_missing)

    # -------------------- asset header fill --------------------

    def _load_asset(self) -> None:
        self._asset_row = _read_asset_row(self.asset_uid) or {}
        if not self._asset_row:
            raise PermissionError("Sredstvo nije pronađeno ili nije dostupno u scope-u.")

        if (not str(self._asset_row.get("nomenclature_no") or "").strip()) and str(self._asset_row.get("nomenclature_number") or "").strip():
            self._asset_row["nomenclature_no"] = self._asset_row.get("nomenclature_number")

        self.lb_uid.setText(self.asset_uid)
        self._set_text_or_dash(self.lb_name, self._asset_row.get("name", ""))
        self._set_text_or_dash(self.lb_status, self._asset_row.get("status", ""))
        self._set_text_or_dash(self.lb_cat, self._asset_row.get("category", ""))
        self._set_text_or_dash(self.lb_toc, self._asset_row.get("toc_number", ""))
        self._set_text_or_dash(self.lb_sn, self._asset_row.get("serial_number", ""))
        self._set_text_or_dash(self.lb_nom, self._asset_row.get("nomenclature_no", ""))
        self._set_text_or_dash(self.lb_inventory, self._asset_row.get("inventory_no", ""))
        self._set_text_or_dash(self.lb_holder, self._asset_row.get("current_holder", ""))
        self._set_text_or_dash(self.lb_loc, self._asset_row.get("location", ""))

        self.lb_created.setText(fmt_date_sr(str(self._asset_row.get("created_at", "") or "")))
        self.lb_updated.setText(fmt_date_sr(str(self._asset_row.get("updated_at", "") or "")))

        self._set_text_or_dash(self.lb_sector, self._asset_row.get("sector", ""))
        self.lb_is_metro.setText("DA" if _safe_int(self._asset_row.get("is_metrology", 0), 0) == 1 else "NE")

        # details mirror
        self.det_uid.setText(self.asset_uid)
        self._set_text_or_dash(self.det_name, self._asset_row.get("name", ""))
        self._set_text_or_dash(self.det_status, self._asset_row.get("status", ""))
        self._set_text_or_dash(self.det_cat, self._asset_row.get("category", ""))
        self._set_text_or_dash(self.det_toc, self._asset_row.get("toc_number", ""))
        self._set_text_or_dash(self.det_sn, self._asset_row.get("serial_number", ""))
        self._set_text_or_dash(self.det_nom, self._asset_row.get("nomenclature_no", ""))
        self._set_text_or_dash(self.det_holder, self._asset_row.get("current_holder", ""))
        self._set_text_or_dash(self.det_loc, self._asset_row.get("location", ""))
        self._set_text_or_dash(self.det_sector, self._asset_row.get("sector", ""))
        self.det_is_metro.setText("DA" if _safe_int(self._asset_row.get("is_metrology", 0), 0) == 1 else "NE")

        self.det_created.setText(fmt_date_sr(str(self._asset_row.get("created_at", "") or "")))
        self.det_updated.setText(fmt_date_sr(str(self._asset_row.get("updated_at", "") or "")))
        self._set_text_or_dash(self.det_vendor, self._asset_row.get("vendor", ""))
        self._set_text_or_dash(self.det_model, self._asset_row.get("model", ""))
        self._set_text_or_dash(self.det_inventory, self._asset_row.get("inventory_no", ""))

        try:
            self.det_notes.setPlainText(str(self._asset_row.get("notes", "") or ""))
        except Exception:
            pass

        self._apply_missing_warnings()

        try:
            self.btn_edit.setEnabled(_can_asset_edit() and (AssetEditDialog is not None))
            if AssetEditDialog is None and _can_asset_edit():
                self.btn_edit.setToolTip(f"Izmena nije dostupna (nema ui/asset_edit_dialog.py). {_EDIT_IMPORT_ERR}")
        except Exception:
            pass

        try:
            self.btn_scrap.setEnabled(_can_asset_edit() or _can_disposal_prepare() or _can_disposal_dispose())
        except Exception:
            pass

    # -------------------- disposal badge + menu --------------------

    def _get_open_disposal_case_db(self) -> Optional[Dict[str, Any]]:
        db_path = _resolve_db_path()
        if not db_path.exists():
            return None
        try:
            conn = sqlite3.connect(db_path.as_posix())
            conn.row_factory = sqlite3.Row
        except Exception:
            return None
        try:
            t = conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='disposal_cases' LIMIT 1;").fetchone()
            if not t:
                return None
            row = conn.execute(
                """
                SELECT disposal_id, created_at, asset_uid, status, prepared_by, reason, notes, approved_by, approved_at,
                       disposed_by, disposed_at, disposed_doc_no, source
                FROM disposal_cases
                WHERE asset_uid=? AND status IN ('PREPARED','APPROVED')
                ORDER BY created_at DESC, disposal_id DESC
                LIMIT 1;
                """,
                (self.asset_uid,),
            ).fetchone()
            if not row:
                return None
            return {k: row[k] for k in row.keys()}
        except Exception:
            return None
        finally:
            try:
                conn.close()
            except Exception:
                pass

    def _refresh_disposal_badge(self) -> None:
        st_asset = str(self._asset_row.get("status") or "").strip().lower()
        if st_asset == "scrapped":
            self.lb_disposal.setText("ZAVRŠENO: sredstvo je rashodovano (status=scrapped).")
            self.lb_disposal.setStyleSheet("color:#e67e22; font-weight:700;")
            return

        case = self._get_open_disposal_case_db()
        if not case:
            self.lb_disposal.setText("Nema otvorene pripreme za rashod.")
            self.lb_disposal.setStyleSheet("color:#999; font-weight:600;")
            return

        st = str(case.get("status") or "").strip().upper()
        did = int(case.get("disposal_id") or 0)

        if st == "PREPARED":
            self.lb_disposal.setText(f"OTVORENO: PRIPREMLJENO (case #{did}) — čeka odobrenje.")
            self.lb_disposal.setStyleSheet("color:#f1c40f; font-weight:700;")
        elif st == "APPROVED":
            self.lb_disposal.setText(f"OTVORENO: ODOBRENO (case #{did}) — spremno za finalni rashod.")
            self.lb_disposal.setStyleSheet("color:#2ecc71; font-weight:700;")
        else:
            self.lb_disposal.setText(f"Otvoreno: {st} (case #{did})")
            self.lb_disposal.setStyleSheet("color:#999; font-weight:600;")

    def _open_disposal_menu(self) -> None:
        m = QMenu(self)

        st_asset = str(self._asset_row.get("status") or "").strip().lower()
        case = self._get_open_disposal_case_db()

        if st_asset == "scrapped":
            a_restore = m.addAction("Vrati u aktivno (legacy)")
            a_restore.setEnabled(_can_asset_edit())
            act = m.exec(self.btn_scrap.mapToGlobal(self.btn_scrap.rect().bottomLeft()))
            if act == a_restore:
                self._legacy_restore_active()
            return

        a_prepare = m.addAction("Priprema za rashod…")
        a_prepare.setEnabled(_can_disposal_prepare())

        a_approve = m.addAction("Odobri rashod")
        a_approve.setEnabled(False)

        a_dispose = m.addAction("Rashoduj (final)")
        a_dispose.setEnabled(False)

        a_cancel = m.addAction("Otkaži pripremu")
        a_cancel.setEnabled(False)

        m.addSeparator()

        a_legacy_scrap = m.addAction("Rashoduj odmah (legacy)")
        a_legacy_scrap.setToolTip("Direktno postavlja status sredstva na 'scrapped' (bez disposal case-a).")
        a_legacy_scrap.setEnabled(_can_asset_edit() and _can_disposal_dispose())

        if case:
            st = str(case.get("status") or "").strip().upper()
            did = int(case.get("disposal_id") or 0)
            if st == "PREPARED":
                a_approve.setEnabled(_can_disposal_approve())
                a_cancel.setEnabled(_can_disposal_prepare())
                a_prepare.setEnabled(False)
                a_prepare.setText(f"Priprema… (case #{did} već postoji)")
            elif st == "APPROVED":
                a_dispose.setEnabled(_can_disposal_dispose())
                a_cancel.setEnabled(_can_disposal_prepare())
                a_prepare.setEnabled(False)
                a_prepare.setText(f"Priprema… (case #{did} već postoji)")

        act = m.exec(self.btn_scrap.mapToGlobal(self.btn_scrap.rect().bottomLeft()))
        if act is None:
            return

        if act == a_prepare:
            self._ui_prepare_disposal()
        elif act == a_approve:
            self._ui_approve_disposal(case)
        elif act == a_dispose:
            self._ui_dispose_from_case(case)
        elif act == a_cancel:
            self._ui_cancel_disposal(case)
        elif act == a_legacy_scrap:
            self._legacy_scrap()

# (FILENAME: ui/asset_detail_dialog.py - END PART 3/4)

# FILENAME: ui/asset_detail_dialog.py
# (FILENAME: ui/asset_detail_dialog.py - START PART 4/4)

    # -------------------- disposal actions --------------------

    def _ui_prepare_disposal(self) -> None:
        if not _can_disposal_prepare():
            QMessageBox.warning(self, "RBAC", "Nemaš pravo: disposal.prepare")
            return

        reason, ok = QInputDialog.getText(self, "Priprema za rashod", "Razlog (kratko):")
        if not ok:
            return
        notes, ok2 = QInputDialog.getMultiLineText(self, "Priprema za rashod", "Napomena (opciono):", "")
        if not ok2:
            return

        try:
            if svc_prepare_disposal is None:
                raise RuntimeError("prepare_disposal nije dostupan (services/assets_service.py).")
            _ = svc_prepare_disposal(  # type: ignore[misc]
                asset_uid=self.asset_uid,
                reason=(reason or "").strip(),
                notes=(notes or "").strip(),
                data=None,
                source="ui_asset_detail.prepare_disposal",
            )
            QMessageBox.information(self, "OK", "Priprema je otvorena (PREPARED).")
            self._reload_all()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da pripremim rashod.\n\n{e}")

    def _ui_approve_disposal(self, case: Optional[Dict[str, Any]]) -> None:
        if not case:
            QMessageBox.information(self, "Info", "Nema otvorenog slučaja.")
            return
        if not _can_disposal_approve():
            QMessageBox.warning(self, "RBAC", "Nemaš pravo: disposal.approve")
            return

        did = int(case.get("disposal_id") or 0)
        if did <= 0:
            QMessageBox.warning(self, "Greška", "Neispravan disposal_id.")
            return

        reply = QMessageBox.question(self, "Potvrda", f"Odobriti rashod? (case #{did})", QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            if svc_approve_disposal is None:
                raise RuntimeError("approve_disposal nije dostupan (services/assets_service.py).")
            svc_approve_disposal(disposal_id=did, source="ui_asset_detail.approve_disposal")  # type: ignore[misc]
            QMessageBox.information(self, "OK", "Rashod odobren (APPROVED).")
            self._reload_all()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da odobrim.\n\n{e}")

    def _ui_cancel_disposal(self, case: Optional[Dict[str, Any]]) -> None:
        if not case:
            QMessageBox.information(self, "Info", "Nema otvorenog slučaja.")
            return
        if not _can_disposal_prepare():
            QMessageBox.warning(self, "RBAC", "Nemaš pravo: disposal.prepare")
            return

        did = int(case.get("disposal_id") or 0)
        if did <= 0:
            QMessageBox.warning(self, "Greška", "Neispravan disposal_id.")
            return

        why, ok = QInputDialog.getText(self, "Otkaži pripremu", "Razlog otkaza (opciono):")
        if not ok:
            return

        reply = QMessageBox.question(self, "Potvrda", f"Otkazati pripremu? (case #{did})", QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            if svc_cancel_disposal is None:
                raise RuntimeError("cancel_disposal nije dostupan (services/assets_service.py).")
            svc_cancel_disposal(disposal_id=did, reason=(why or "").strip(), source="ui_asset_detail.cancel_disposal")  # type: ignore[misc]
            QMessageBox.information(self, "OK", "Priprema otkazana (CANCELLED).")
            self._reload_all()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da otkažem.\n\n{e}")

    def _ui_dispose_from_case(self, case: Optional[Dict[str, Any]]) -> None:
        if not case:
            QMessageBox.information(self, "Info", "Nema otvorenog slučaja.")
            return
        if not _can_disposal_dispose():
            QMessageBox.warning(self, "RBAC", "Nemaš pravo: disposal.dispose")
            return

        did = int(case.get("disposal_id") or 0)
        if did <= 0:
            QMessageBox.warning(self, "Greška", "Neispravan disposal_id.")
            return

        doc_no, ok = QInputDialog.getText(self, "Final rashod", "Broj dokumenta (opciono):")
        if not ok:
            return

        reply = QMessageBox.question(
            self,
            "Potvrda",
            f"Finalno rashodovati sredstvo? (case #{did})\n\nOvo postavlja status na 'scrapped'.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        try:
            if svc_dispose_from_case is None:
                raise RuntimeError("dispose_from_case nije dostupan (services/assets_service.py).")
            svc_dispose_from_case(  # type: ignore[misc]
                disposal_id=did,
                disposed_doc_no=(doc_no or "").strip(),
                source="ui_asset_detail.dispose_from_case",
            )
            QMessageBox.information(self, "OK", "Sredstvo rashodovano (DISPOSED + status scrapped).")
            self._reload_all()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da rashodujem.\n\n{e}")

    def _legacy_scrap(self) -> None:
        if not (_can_asset_edit() and _can_disposal_dispose()):
            QMessageBox.warning(self, "RBAC", "Nemaš pravo (assets.edit + disposal.dispose).")
            return
        reply = QMessageBox.question(
            self,
            "Legacy rashod",
            "Rashodovati odmah (legacy)?\n\nDirektno status -> 'scrapped', bez disposal case-a.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        try:
            from core.db import update_asset_db  # type: ignore
            update_asset_db(actor=actor_name(), asset_uid=self.asset_uid, status="scrapped", source="ui_asset_detail_legacy_scrap")
            QMessageBox.information(self, "OK", "Sredstvo rashodovano (legacy).")
            self._reload_all()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da promenim status.\n\n{e}")

    def _legacy_restore_active(self) -> None:
        if not _can_asset_edit():
            QMessageBox.warning(self, "RBAC", "Nemaš pravo (assets.edit).")
            return
        reply = QMessageBox.question(
            self,
            "Povratak",
            "Vratiti sredstvo u aktivno stanje (legacy)?\n\nStatus -> 'active'.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        try:
            from core.db import update_asset_db  # type: ignore
            update_asset_db(actor=actor_name(), asset_uid=self.asset_uid, status="active", source="ui_asset_detail_legacy_restore")
            QMessageBox.information(self, "OK", "Sredstvo vraćeno u aktivno.")
            self._reload_all()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da promenim status.\n\n{e}")

    # -------------------- edit asset --------------------

    def _edit_asset(self) -> None:
        if not _can_asset_edit():
            QMessageBox.warning(self, "Zabranjeno", "Nemaš pravo za izmenu.")
            return

        if AssetEditDialog is None:
            msg = "Edit dijalog nije dostupan (ui/asset_edit_dialog.py nije učitan)."
            if _EDIT_IMPORT_ERR:
                msg += f"\n\nDetalj: {_EDIT_IMPORT_ERR}"
            QMessageBox.information(self, "Izmena nije dostupna", msg)
            return

        try:
            try:
                dlg = AssetEditDialog(self.asset_uid, initial=dict(self._asset_row), parent=self)  # type: ignore[misc]
            except TypeError:
                try:
                    dlg = AssetEditDialog(asset_uid=self.asset_uid, initial=dict(self._asset_row), parent=self)  # type: ignore[misc]
                except TypeError:
                    try:
                        dlg = AssetEditDialog(self.asset_uid, parent=self)  # type: ignore[misc]
                    except TypeError:
                        dlg = AssetEditDialog(asset_uid=self.asset_uid, parent=self)  # type: ignore[misc]

            if dlg.exec() == QDialog.Accepted:
                self._reload_all()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da otvorim izmenu.\n\n{e}")

    # -------------------- timeline --------------------

    def _load_timeline(self, *args) -> None:
        rows, msg = _load_timeline_rows(self.asset_uid, limit=400)
        try:
            self.lb_tl_info.setText(msg)
        except Exception:
            pass

        def _to_display_dt(v: Any) -> str:
            s = str(v or "").strip()
            if not s:
                return ""
            try:
                s2 = s.replace("Z", "").replace("T", " ").strip()
                dt = datetime.fromisoformat(s2)
                return dt.strftime("%d.%m.%Y %H:%M:%S")
            except Exception:
                return s

        try:
            self.tbl_tl.setUpdatesEnabled(False)
            self.tbl_tl.setSortingEnabled(False)
            self.tbl_tl.setRowCount(0)

            keys = list(rows[0].keys()) if rows else []

            def pick(*cands: str) -> str:
                for c in cands:
                    if c in keys:
                        return c
                return ""

            when_key = pick("event_time", "created_at", "ts", "timestamp", "time")
            actor_key = pick("actor", "user", "username", "performed_by", "by_user", "who")
            action_key = pick("event_type", "action", "event", "type", "op")
            detail_key = pick("details", "detail", "message", "payload", "data_json", "data", "note", "changes", "description")

            for rr in rows:
                i = self.tbl_tl.rowCount()
                self.tbl_tl.insertRow(i)

                w = rr.get(when_key, "") if when_key else ""
                a = rr.get(actor_key, "") if actor_key else ""
                ac = rr.get(action_key, "") if action_key else ""
                d = rr.get(detail_key, "") if detail_key else ""

                self.tbl_tl.setItem(i, 0, QTableWidgetItem(_to_display_dt(w)))
                self.tbl_tl.setItem(i, 1, QTableWidgetItem(str(a or "")))
                self.tbl_tl.setItem(i, 2, QTableWidgetItem(str(ac or "")))
                self.tbl_tl.setItem(i, 3, QTableWidgetItem(str(d or "")))

            try:
                self.tbl_tl.resizeColumnsToContents()
            except Exception:
                pass
        except Exception as e:
            try:
                self.lb_tl_info.setText(f"Timeline greška: {e}")
            except Exception:
                pass
        finally:
            try:
                self.tbl_tl.setUpdatesEnabled(True)
                self.tbl_tl.setSortingEnabled(True)
            except Exception:
                pass

    # -------------------- metrology --------------------

    def _can_manage_metrology(self) -> bool:
        try:
            return bool(can(PERM_METRO_MANAGE))
        except Exception:
            return False

    def _can_view_metrology_perm(self) -> bool:
        try:
            return bool(can(PERM_METRO_VIEW))
        except Exception:
            return False

    def _can_sector_scope_metrology(self) -> bool:
        try:
            if not (bool(can(PERM_ASSETS_METRO_VIEW)) and bool(can(PERM_METRO_VIEW))):
                return False
        except Exception:
            return False

        asset_sec = str(self._asset_row.get("sector", "") or "").strip()
        user_sec = str(current_sector() or "").strip()
        is_m = _safe_int(self._asset_row.get("is_metrology", 0), 0) == 1
        return bool(is_m and asset_sec and user_sec and (_norm(asset_sec) == _norm(user_sec)))

    def _apply_metrology_scope_ui(self) -> None:
        if list_metrology_records_for_asset is None:
            self.cb_met_warn.setEnabled(False)
            self.btn_met_refresh.setEnabled(False)
            self.tbl_met.setEnabled(False)
            self.btn_met_open.setEnabled(False)
            msg = "Metrologija modul nije dostupan."
            if _METRO_IMPORT_ERR:
                msg += f" ({_METRO_IMPORT_ERR})"
            self.lb_met_info.setText(msg)
            return

        if not self._can_view_metrology_perm():
            self.cb_met_warn.setEnabled(False)
            self.btn_met_refresh.setEnabled(False)
            self.tbl_met.setEnabled(False)
            self.btn_met_open.setEnabled(False)
            self.lb_met_info.setText("Nemaš pravo (metrology.view).")
            return

        self.cb_met_warn.setEnabled(True)
        self.btn_met_refresh.setEnabled(True)
        self.tbl_met.setEnabled(True)

        is_m = _safe_int(self._asset_row.get("is_metrology", 0), 0) == 1
        if self._can_manage_metrology():
            base = "Metrologija: MANAGE (global)."
        elif self._can_sector_scope_metrology():
            base = "Metrologija: sektor-scope."
        else:
            base = "Metrologija: scope enforced u servisu."

        if not is_m:
            base += " | Napomena: flag=NE."

        self.lb_met_info.setText(base)

    def _sync_met_buttons(self) -> None:
        try:
            self.btn_met_open.setEnabled(_current_row_any(self.tbl_met) >= 0 and self.tbl_met.isEnabled())
        except Exception:
            pass

    def _selected_met_uid(self) -> str:
        r = _current_row_any(self.tbl_met)
        if r < 0:
            return ""
        it = self.tbl_met.item(r, 1)
        return it.text().strip() if it else ""

    def _load_metrology(self, *args) -> None:
        self._apply_metrology_scope_ui()

        if not self.tbl_met.isEnabled():
            try:
                self.tbl_met.setRowCount(0)
            except Exception:
                pass
            self._sync_met_buttons()
            return

        try:
            warn_days = int(self.cb_met_warn.currentText())
        except Exception:
            warn_days = 30

        rows: List[Dict[str, Any]] = []
        try:
            fn = list_metrology_records_for_asset
            if fn is None:
                raise RuntimeError(_METRO_IMPORT_ERR or "Metrology funkcija nije dostupna.")

            try:
                rows = fn(self.asset_uid, warn_days=warn_days, limit=2000) or []
            except TypeError:
                try:
                    rows = fn(self.asset_uid, warn_days=warn_days) or []
                except TypeError:
                    try:
                        rows = fn(self.asset_uid, limit=2000) or []
                    except TypeError:
                        rows = fn(self.asset_uid) or []

            base = (self.lb_met_info.text() or "").strip()
            self.lb_met_info.setText((base + " | " if base else "") + f"Zapisa: {len(rows)}")
        except PermissionError:
            rows = []
            self.lb_met_info.setText("Nemaš pravo (RBAC scope u servisu).")
        except Exception as e:
            rows = []
            self.lb_met_info.setText(f"Greška: {e}")

        try:
            self.tbl_met.setUpdatesEnabled(False)
            self.tbl_met.setSortingEnabled(False)
            self.tbl_met.setRowCount(0)

            for rr in rows:
                i = self.tbl_met.rowCount()
                self.tbl_met.insertRow(i)
                vals = [
                    rr.get("status", "") or "",
                    rr.get("met_uid", "") or "",
                    rr.get("calib_type", "") or "",
                    fmt_date_sr(str(rr.get("calib_date", "") or "")),
                    fmt_date_sr(str(rr.get("valid_until", "") or "")),
                    rr.get("provider_name", "") or "",
                    rr.get("cert_no", "") or "",
                ]
                for c, v in enumerate(vals):
                    self.tbl_met.setItem(i, c, QTableWidgetItem(str(v)))
        finally:
            try:
                self.tbl_met.setUpdatesEnabled(True)
                self.tbl_met.setSortingEnabled(True)
            except Exception:
                pass

        self._sync_met_buttons()

    def _open_met_details(self) -> None:
        met_uid = self._selected_met_uid()
        if not met_uid:
            return
        try:
            from ui.metrology_page import MetrologyDetailsDialog  # type: ignore
            try:
                warn_days = int(self.cb_met_warn.currentText())
            except Exception:
                warn_days = 30
            dlg = MetrologyDetailsDialog(met_uid, parent=self, warn_days=warn_days)
            dlg.exec()
        except PermissionError as e:
            QMessageBox.information(self, "RBAC", f"Nemaš pravo.\n\n{e}")
        except Exception as e:
            QMessageBox.information(self, "Metrologija", f"Ne mogu da otvorim detalje.\n\n{e}")

    def _on_met_context_menu(self, pos) -> None:
        if not self.tbl_met.isEnabled():
            return
        r = _current_row_any(self.tbl_met)
        if r < 0:
            return

        met_uid = self.tbl_met.item(r, 1).text() if self.tbl_met.item(r, 1) else ""
        m = QMenu(self)
        a_copy_uid = m.addAction("Kopiraj Met UID")
        a_copy_row = m.addAction("Kopiraj selekciju (Ctrl+C)")
        m.addSeparator()
        a_open = m.addAction("Detalji zapisa")

        act = m.exec(self.tbl_met.mapToGlobal(pos))
        if act == a_copy_uid:
            _set_clipboard_text(str(met_uid or ""))
        elif act == a_copy_row:
            try:
                copy_selected_cells(self.tbl_met)
            except Exception:
                pass
        elif act == a_open:
            self._open_met_details()

    # -------------------- calendar --------------------

    def _apply_calendar_scope_ui(self) -> None:
        if list_calendar_events_for_asset is None:
            self.btn_cal_refresh.setEnabled(False)
            self.btn_cal_add.setEnabled(False)
            self.btn_cal_edit.setEnabled(False)
            self.btn_cal_del.setEnabled(False)
            try:
                self.cal.setEnabled(False)
                self.tbl_cal.setEnabled(False)
            except Exception:
                pass
            msg = "Kalendar modul nije dostupan."
            if _CAL_IMPORT_ERR:
                msg += f" ({_CAL_IMPORT_ERR})"
            self.lb_cal_info.setText(msg)
            return

        self.btn_cal_refresh.setEnabled(True)
        self.btn_cal_add.setEnabled(True)
        try:
            self.cal.setEnabled(True)
            self.tbl_cal.setEnabled(True)
        except Exception:
            pass
        if (self.lb_cal_info.text() or "").startswith("Kalendar modul nije dostupan"):
            self.lb_cal_info.setText("")

    def _sync_calendar_buttons(self) -> None:
        r = _current_row_any(self.tbl_cal)
        has = r >= 0 and self.tbl_cal.isEnabled()
        self.btn_cal_edit.setEnabled(has and (update_calendar_event is not None))
        self.btn_cal_del.setEnabled(has and (delete_calendar_event is not None))

        self._cal_selected_id = 0
        if has:
            it = self.tbl_cal.item(r, 0)
            try:
                v = it.data(Qt.UserRole) if it is not None else None
                if v is None:
                    v = it.text() if it is not None else "0"
                self._cal_selected_id = int(v)
            except Exception:
                self._cal_selected_id = 0

    def _selected_calendar_date_iso(self) -> str:
        d = self.cal.selectedDate()
        return f"{d.year():04d}-{d.month():02d}-{d.day():02d}"

    def _is_valid_time_hhmm(self, s: str) -> bool:
        s = (s or "").strip()
        if not s:
            return True
        if len(s) != 5 or s[2] != ":":
            return False
        try:
            hh = int(s[0:2])
            mm = int(s[3:5])
            return 0 <= hh <= 23 and 0 <= mm <= 59
        except Exception:
            return False

    def _load_calendar_for_selected_date(self, *args) -> None:
        self._apply_calendar_scope_ui()

        try:
            self.tbl_cal.setRowCount(0)
        except Exception:
            pass

        if list_calendar_events_for_asset is None:
            self.lb_cal_info.setText(self.lb_cal_info.text() or "Kalendar modul nije dostupan.")
            return

        day_iso = self._selected_calendar_date_iso()

        try:
            rows = list_calendar_events_for_asset(self.asset_uid, day_iso) or []  # type: ignore[misc]
        except TypeError:
            rows = list_calendar_events_for_asset(asset_uid=self.asset_uid, day_iso=day_iso) or []  # type: ignore[misc]
        except Exception as e:
            self.lb_cal_info.setText(f"Kalendar greška: {e}")
            rows = []

        try:
            self.tbl_cal.setUpdatesEnabled(False)
            self.tbl_cal.setSortingEnabled(False)
            self.tbl_cal.setRowCount(0)

            for rr in rows:
                i = self.tbl_cal.rowCount()
                self.tbl_cal.insertRow(i)

                eid = rr.get("id", rr.get("event_id", 0)) or 0
                when = rr.get("time", rr.get("when", rr.get("starts_at", ""))) or ""
                etype = rr.get("event_type", rr.get("type", "")) or ""
                text = rr.get("text", rr.get("note", rr.get("title", ""))) or ""

                it0 = QTableWidgetItem(str(eid))
                try:
                    it0.setData(Qt.UserRole, int(eid))
                except Exception:
                    pass

                self.tbl_cal.setItem(i, 0, it0)
                self.tbl_cal.setItem(i, 1, QTableWidgetItem(str(when)))
                self.tbl_cal.setItem(i, 2, QTableWidgetItem(str(etype)))
                self.tbl_cal.setItem(i, 3, QTableWidgetItem(str(text)))
        finally:
            try:
                self.tbl_cal.setUpdatesEnabled(True)
                self.tbl_cal.setSortingEnabled(True)
            except Exception:
                pass

        self.lb_cal_info.setText(f"{day_iso} — događaja: {self.tbl_cal.rowCount()}")
        self._sync_calendar_buttons()

    def _calendar_add(self, *args) -> None:
        if add_calendar_event_for_asset is None:
            QMessageBox.information(self, "Info", "Kalendar servis nije dostupan (nema add funkcije).")
            return

        day_iso = self._selected_calendar_date_iso()

        try:
            txt, ok = QInputDialog.getText(self, "Novi događaj", "Opis (npr. 'Servis', 'Etaloniranje', 'Napomena'):")
            if not ok or not (txt or "").strip():
                return

            etype, ok2 = QInputDialog.getText(self, "Tip", "Tip (npr. SERVICE / METRO / NOTE / TRANSFER):")
            if not ok2:
                return
            etype = (etype or "").strip() or "NOTE"

            tm, ok3 = QInputDialog.getText(self, "Vreme", "Vreme (HH:MM) (opciono):")
            if not ok3:
                return
            tm = (tm or "").strip()
            if not self._is_valid_time_hhmm(tm):
                QMessageBox.warning(self, "Vreme", "Neispravan format. Koristi HH:MM ili ostavi prazno.")
                return

            add_calendar_event_for_asset(  # type: ignore[misc]
                actor=actor_name(),
                asset_uid=self.asset_uid,
                day_iso=day_iso,
                time_hhmm=tm,
                event_type=etype,
                text=txt.strip(),
                source="ui_asset_detail_calendar_add",
            )
            self._load_calendar_for_selected_date()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da dodam događaj.\n\n{e}")

    def _calendar_edit(self, *args) -> None:
        if update_calendar_event is None:
            QMessageBox.information(self, "Info", "Kalendar servis nema edit funkciju.")
            return
        if not self._cal_selected_id:
            QMessageBox.information(self, "Info", "Prvo izaberi događaj u tabeli.")
            return

        r = _current_row_any(self.tbl_cal)
        if r < 0:
            return

        cur_time = self.tbl_cal.item(r, 1).text() if self.tbl_cal.item(r, 1) else ""
        cur_type = self.tbl_cal.item(r, 2).text() if self.tbl_cal.item(r, 2) else ""
        cur_txt = self.tbl_cal.item(r, 3).text() if self.tbl_cal.item(r, 3) else ""

        try:
            txt, ok = QInputDialog.getText(self, "Izmeni događaj", "Opis:", text=cur_txt)
            if not ok:
                return
            etype, ok2 = QInputDialog.getText(self, "Izmeni tip", "Tip:", text=cur_type)
            if not ok2:
                return
            tm, ok3 = QInputDialog.getText(self, "Izmeni vreme", "Vreme (HH:MM):", text=cur_time)
            if not ok3:
                return

            tm = (tm or "").strip()
            if not self._is_valid_time_hhmm(tm):
                QMessageBox.warning(self, "Vreme", "Neispravan format. Koristi HH:MM ili prazno.")
                return

            update_calendar_event(  # type: ignore[misc]
                actor=actor_name(),
                event_id=int(self._cal_selected_id),
                time_hhmm=tm,
                event_type=(etype or "").strip(),
                text=(txt or "").strip(),
                source="ui_asset_detail_calendar_edit",
            )
            self._load_calendar_for_selected_date()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da izmenim događaj.\n\n{e}")

    def _calendar_delete(self, *args) -> None:
        if delete_calendar_event is None:
            QMessageBox.information(self, "Info", "Kalendar servis nema delete funkciju.")
            return
        if not self._cal_selected_id:
            QMessageBox.information(self, "Info", "Prvo izaberi događaj u tabeli.")
            return

        reply = QMessageBox.question(self, "Potvrda brisanja", "Obrisati događaj?", QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            delete_calendar_event(  # type: ignore[misc]
                actor=actor_name(),
                event_id=int(self._cal_selected_id),
                source="ui_asset_detail_calendar_delete",
            )
            self._load_calendar_for_selected_date()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da obrišem događaj.\n\n{e}")

    # -------------------- custom fields --------------------

    def _clear_custom_form(self) -> None:
        try:
            while self.custom_form.rowCount() > 1:
                item_label = self.custom_form.itemAt(0, QFormLayout.LabelRole)
                item_field = self.custom_form.itemAt(0, QFormLayout.FieldRole)
                w1 = item_label.widget() if item_label else None
                w2 = item_field.widget() if item_field else None
                self.custom_form.removeRow(0)
                if w1:
                    w1.deleteLater()
                if w2:
                    w2.deleteLater()
        except Exception:
            pass

    def _load_custom_fields(self) -> None:
        self._clear_custom_form()
        self._custom_defs = []
        self._custom_widgets = {}

        try:
            defs = list_field_defs() or []
        except Exception as e:
            defs = []
            QMessageBox.warning(self, "Dodatna polja", f"Ne mogu da učitam definicije.\n\n{e}")

        try:
            vals = list_values_for_asset(self.asset_uid) or []
        except Exception:
            vals = []

        val_map: Dict[str, Any] = {}
        for v in vals:
            k = str(v.get("field_key", "") or v.get("key", "") or "").strip()
            if not k:
                continue
            val_map[k] = v.get("value", v.get("val", ""))

        self._custom_defs = defs

        for d in defs:
            key = str(d.get("field_key", d.get("key", "")) or "").strip()
            label = str(d.get("label", d.get("field_label", key)) or key).strip()
            ftype = str(d.get("field_type", d.get("type", "text")) or "text").strip().lower()
            if not key:
                continue

            cur = val_map.get(key, "")
            if ftype in ("bool", "boolean", "check"):
                w = QCheckBox()
                try:
                    w.setChecked(str(cur).strip().lower() in ("1", "true", "da", "yes", "y"))
                except Exception:
                    pass
            else:
                w = QLineEdit()
                w.setText("" if cur is None else str(cur))

            self._custom_widgets[key] = w
            self.custom_form.insertRow(self.custom_form.rowCount() - 1, label, w)

    def _save_custom_fields(self, *args) -> None:
        payload: Dict[str, Any] = {}
        for k, w in self._custom_widgets.items():
            if isinstance(w, QCheckBox):
                payload[k] = "1" if w.isChecked() else "0"
            else:
                try:
                    payload[k] = w.text().strip()
                except Exception:
                    payload[k] = ""

        try:
            try:
                bulk_set_values_for_asset(
                    actor=actor_name(),
                    asset_uid=self.asset_uid,
                    values=payload,
                    source="ui_asset_detail_custom_save",
                )
            except TypeError:
                bulk_set_values_for_asset(self.asset_uid, payload)  # type: ignore[misc]
            QMessageBox.information(self, "OK", "Dodatna polja su sačuvana.")
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da sačuvam dodatna polja.\n\n{e}")

    # ============================
    # ATTACHMENTS (FIXED + FULL)
    # ============================

    def _selected_attachment_row(self) -> Optional[Dict[str, Any]]:
        try:
            r = _current_row_any(self.tbl)
            if r < 0:
                return None
            it = self.tbl.item(r, 1)  # Naziv
            if it is None:
                return None
            aid = it.data(Qt.UserRole)
            if aid is None:
                if 0 <= r < len(self._att_rows):
                    return self._att_rows[r]
                return None

            aid_int = _safe_int(aid, -1)
            if aid_int < 0:
                return None

            for rr in self._att_rows:
                rr_id = rr.get("id", rr.get("attachment_id", None))
                if rr_id is not None and _safe_int(rr_id, -999999) == aid_int:
                    return rr
            return None
        except Exception:
            return None

    def _attachment_abs_path(self, row: Dict[str, Any]) -> Optional[Path]:
        aid = row.get("id", row.get("attachment_id", None))
        rel = row.get("path", row.get("rel_path", row.get("filepath", row.get("file_path", ""))))

        try:
            if aid is not None:
                p = get_attachment_abs_path(aid)
                if p:
                    return Path(str(p))
        except Exception:
            pass

        if rel:
            p2 = Path(str(rel))
            if p2.is_absolute():
                return p2
            return _safe_rel_under_root(str(rel))
        return None

    def _load_attachments(self, show_errors: bool = True) -> None:
        prev_sel_id: Optional[int] = None
        try:
            cur = self._selected_attachment_row()
            if cur:
                rr_id = cur.get("id", cur.get("attachment_id", None))
                if rr_id is not None:
                    prev_sel_id = int(rr_id)
        except Exception:
            prev_sel_id = None

        try:
            rows = list_attachments_for_asset(self.asset_uid) or []
        except Exception as e:
            rows = []
            if show_errors:
                QMessageBox.warning(self, "Prilozi", f"Ne mogu da učitam priloge.\n\n{e}")

        self._att_rows = rows
        self._att_selected = None
        self._preview_img_path = None

        try:
            self.tbl.setUpdatesEnabled(False)
            self.tbl.setSortingEnabled(False)
            self.tbl.setRowCount(0)

            for idx, rr in enumerate(rows, start=1):
                i = self.tbl.rowCount()
                self.tbl.insertRow(i)

                name = rr.get("filename", rr.get("name", "")) or ""
                created = rr.get("created_at", rr.get("created", "")) or ""
                note = rr.get("note", rr.get("comment", "")) or ""

                rr_id = rr.get("id", rr.get("attachment_id", None))
                rr_id_int = _safe_int(rr_id, 0)

                it0 = QTableWidgetItem(str(idx))
                it1 = QTableWidgetItem(str(name))
                it2 = QTableWidgetItem(fmt_date_sr(str(created or "")))
                it3 = QTableWidgetItem(str(note))
                it1.setData(Qt.UserRole, rr_id_int)  # sort-safety

                self.tbl.setItem(i, 0, it0)
                self.tbl.setItem(i, 1, it1)
                self.tbl.setItem(i, 2, it2)
                self.tbl.setItem(i, 3, it3)

        finally:
            try:
                self.tbl.setUpdatesEnabled(True)
                self.tbl.setSortingEnabled(True)
            except Exception:
                pass

        # restore selection
        if prev_sel_id is not None:
            try:
                for r in range(self.tbl.rowCount()):
                    it = self.tbl.item(r, 1)
                    if it and _safe_int(it.data(Qt.UserRole), -1) == prev_sel_id:
                        self.tbl.setCurrentCell(r, 1)
                        break
            except Exception:
                pass

        try:
            self.ed_att_path.setText("")
            self.btn_copy_path.setEnabled(False)
        except Exception:
            pass

        self._sync_attachment_buttons()
        self._show_preview_none()

    def _sync_attachment_buttons(self) -> None:
        sel = self._selected_attachment_row()
        has = sel is not None
        self.btn_open.setEnabled(has)
        self.btn_open_folder.setEnabled(has)
        self.btn_del.setEnabled(has)
        try:
            self.btn_copy_path.setEnabled(bool(has and (self.ed_att_path.text() or "").strip()))
        except Exception:
            pass

    def _show_preview_none(self) -> None:
        try:
            self.preview_stack.setCurrentIndex(0)
        except Exception:
            pass
        try:
            self.preview_img.setText("Nema pregleda.")
            self.preview_img.setPixmap(QPixmap())
        except Exception:
            pass
        try:
            self.preview_txt.setPlainText("")
        except Exception:
            pass
        try:
            self.lb_find.setText("")
        except Exception:
            pass

        self._last_find_cursor = None
        self._last_find_text = ""
        self._preview_img_path = None
        self._sync_find_enabled()

    def _on_att_selection_changed(self) -> None:
        self._att_selected = self._selected_attachment_row()

        try:
            if not self._att_selected:
                self.ed_att_path.setText("")
                self.btn_copy_path.setEnabled(False)
            else:
                p = self._attachment_abs_path(self._att_selected)
                txt = str(p) if p else ""
                self.ed_att_path.setText(txt)
                self.btn_copy_path.setEnabled(bool(txt))
        except Exception:
            pass

        self._sync_attachment_buttons()

        try:
            if self._att_preview_timer is not None:
                self._att_preview_timer.stop()
                self._att_preview_timer.start(120)
            else:
                self._render_selected_attachment_preview()
        except Exception:
            self._render_selected_attachment_preview()

    def _render_image_preview(self, p: Path) -> None:
        try:
            self.preview_stack.setCurrentIndex(1)
        except Exception:
            return

        pm = QPixmap(str(p))
        if pm.isNull():
            try:
                self.preview_img.setText("Ne mogu da učitam sliku.")
                self.preview_img.setPixmap(QPixmap())
            except Exception:
                pass
            self._preview_img_path = None
            self._sync_find_enabled()
            return

        try:
            w = max(200, self.preview_img.width())
            h = max(160, self.preview_img.height())
            self.preview_img.setPixmap(pm.scaled(w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        except Exception:
            pass

        self._preview_img_path = p
        self._sync_find_enabled()

    def _get_preview_cache(self) -> Dict[str, Any]:
        try:
            if not hasattr(self, "_preview_cache"):
                self._preview_cache = {}  # type: ignore[attr-defined]
            return self._preview_cache  # type: ignore[attr-defined]
        except Exception:
            return {}

    def _cache_get(self, p: Path) -> Optional[str]:
        try:
            key = str(p)
            st = p.stat()
            cache = self._get_preview_cache()
            rec = cache.get(key)
            if not rec:
                return None
            mtime, size, text = rec
            if float(mtime) == float(st.st_mtime) and int(size) == int(st.st_size):
                return str(text)
            return None
        except Exception:
            return None

    def _cache_put(self, p: Path, text: str) -> None:
        try:
            key = str(p)
            st = p.stat()
            cache = self._get_preview_cache()
            cache[key] = (float(st.st_mtime), int(st.st_size), str(text))
            if len(cache) > 8:
                try:
                    first_key = next(iter(cache.keys()))
                    if first_key != key:
                        cache.pop(first_key, None)
                except Exception:
                    pass
        except Exception:
            pass

    def _stop_preview_thread(self) -> None:
        try:
            if self._preview_thread is not None:
                try:
                    self._preview_thread.requestInterruption()
                except Exception:
                    pass
                self._preview_thread.quit()
                self._preview_thread.wait(250)
        except Exception:
            pass
        self._preview_thread = None
        self._preview_worker = None

    def _start_preview_thread(self, path: Path) -> None:
        self._stop_preview_thread()

        self._preview_job_id += 1
        job_id = self._preview_job_id
        self._preview_job_path = str(path)

        worker = _PreviewWorker(job_id, str(path))
        thread = QThread(self)

        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(self._on_preview_ready)

        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        self._preview_worker = worker
        self._preview_thread = thread
        thread.start()

    def _on_preview_ready(self, job_id: int, path_str: str, text: str) -> None:
        if self._closing:
            return
        try:
            if int(job_id) != int(self._preview_job_id):
                return
            if str(path_str) != str(self._preview_job_path):
                return
        except Exception:
            return

        p = Path(path_str)
        try:
            self.preview_stack.setCurrentIndex(2)
        except Exception:
            pass

        try:
            self.preview_txt.setPlainText(text)
            self.preview_txt.moveCursor(QTextCursor.Start)
        except Exception:
            pass

        self._cache_put(p, text)
        self._sync_find_enabled()

    def _render_selected_attachment_preview(self) -> None:
        row = self._att_selected
        if not row:
            self._show_preview_none()
            return

        p = self._attachment_abs_path(row)
        if not p or not p.exists() or not p.is_file():
            try:
                self.preview_stack.setCurrentIndex(2)
                self.preview_txt.setPlainText("Fajl ne postoji na disku (ili putanja nije dostupna).")
                self.preview_txt.moveCursor(QTextCursor.Start)
            except Exception:
                pass
            self._sync_find_enabled()
            self._preview_img_path = None
            return

        self._last_find_cursor = None
        self._last_find_text = ""
        try:
            self.lb_find.setText("")
        except Exception:
            pass
        self._preview_img_path = None

        if _is_image_file(p):
            self._stop_preview_thread()
            self._render_image_preview(p)
            try:
                self.preview_txt.setPlainText("")
            except Exception:
                pass
            return

        try:
            self.preview_stack.setCurrentIndex(2)
        except Exception:
            pass

        cached = self._cache_get(p)
        if cached is not None:
            self._stop_preview_thread()
            try:
                self.preview_txt.setPlainText(cached)
                self.preview_txt.moveCursor(QTextCursor.Start)
            except Exception:
                pass
            self._sync_find_enabled()
            return

        try:
            self.preview_txt.setPlainText("Učitavam preview…")
            self.preview_txt.moveCursor(QTextCursor.Start)
        except Exception:
            pass

        self._start_preview_thread(p)

    def add_attachment(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Izaberi fajl",
            "",
            "Svi fajlovi (*.*);;Dokumenti (*.pdf *.docx *.xlsx *.txt *.csv *.log);;Slike (*.png *.jpg *.jpeg *.bmp *.gif *.webp)",
        )
        if not file_path:
            return

        p = Path(file_path)
        if not p.exists() or not p.is_file():
            QMessageBox.warning(self, "Prilozi", "Izabrani fajl ne postoji ili nije fajl.")
            return

        try:
            note = (self.ed_note.toPlainText() or "").strip()
        except Exception:
            note = ""

        try:
            size_mb = p.stat().st_size / (1024 * 1024)
            if size_mb > 250:
                reply = QMessageBox.question(
                    self,
                    "Veliki fajl",
                    f"Fajl je velik (~{size_mb:.1f} MB). Dodavanje i preview mogu biti sporiji.\n\nNastaviti?",
                    QMessageBox.Yes | QMessageBox.No,
                )
                if reply != QMessageBox.Yes:
                    return
        except Exception:
            pass

        try:
            try:
                add_attachment_to_asset(
                    actor=actor_name(),
                    asset_uid=self.asset_uid,
                    file_path=file_path,
                    note=note,
                    source="ui_asset_detail_add_attachment",
                )
            except TypeError:
                try:
                    add_attachment_to_asset(self.asset_uid, file_path, note)  # type: ignore[misc]
                except TypeError:
                    add_attachment_to_asset(self.asset_uid, file_path)  # type: ignore[misc]

            try:
                self.ed_note.setPlainText("")
            except Exception:
                pass

            self._load_attachments(show_errors=True)
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da dodam prilog.\n\n{e}")

    def open_attachment(self) -> None:
        row = self._selected_attachment_row()
        if not row:
            return
        p = self._attachment_abs_path(row)
        if not p or not p.exists():
            QMessageBox.warning(self, "Prilozi", "Fajl ne postoji na disku.")
            return

        ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(p)))
        if not ok:
            QMessageBox.warning(self, "Prilozi", "Ne mogu da otvorim fajl (nema default aplikacije ili je blokirano).")

    def open_attachment_folder(self) -> None:
        row = self._selected_attachment_row()
        if not row:
            return
        p = self._attachment_abs_path(row)
        if not p or not p.exists():
            QMessageBox.warning(self, "Prilozi", "Fajl ne postoji na disku.")
            return
        folder = p.parent
        if not folder.exists():
            QMessageBox.warning(self, "Prilozi", "Folder ne postoji na disku.")
            return

        ok = QDesktopServices.openUrl(QUrl.fromLocalFile(str(folder)))
        if not ok:
            QMessageBox.warning(self, "Prilozi", "Ne mogu da otvorim folder (nema file manager-a ili je blokirano).")

    def delete_attachment_ui(self) -> None:
        row = self._selected_attachment_row()
        if not row:
            return

        reply = QMessageBox.question(
            self,
            "Potvrda brisanja",
            "Obrisati izabrani prilog?\n\n(Ovo briše zapis i fajl po pravilima servisa.)",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        aid = row.get("id", row.get("attachment_id", None))
        if aid is None:
            QMessageBox.warning(self, "Prilozi", "Ne mogu da obrišem: nema attachment_id.")
            return

        try:
            try:
                delete_attachment(actor=actor_name(), attachment_id=int(aid), source="ui_asset_detail_delete_attachment")
            except TypeError:
                delete_attachment(int(aid))  # type: ignore[misc]
            self._load_attachments(show_errors=True)
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Ne mogu da obrišem prilog.\n\n{e}")

    # ✅ FIX: metod koji ti je falio (i koji je rušio dijalog)
    def _on_tbl_context_menu(self, pos) -> None:
        r = _current_row_any(self.tbl)
        if r < 0:
            return
        row = self._selected_attachment_row()
        if not row:
            return

        p = self._attachment_abs_path(row)
        ptxt = str(p) if p else ""
        ftxt = str(p.parent) if (p and p.parent) else ""

        m = QMenu(self)
        a_open = m.addAction("Otvori")
        a_open_folder = m.addAction("Otvori folder")
        m.addSeparator()
        a_copy_path = m.addAction("Kopiraj putanju fajla")
        a_copy_folder = m.addAction("Kopiraj putanju foldera")
        a_copy_sel = m.addAction("Kopiraj selekciju (Ctrl+C)")
        m.addSeparator()
        a_del = m.addAction("Obriši")

        act = m.exec(self.tbl.mapToGlobal(pos))
        if act == a_open:
            self.open_attachment()
        elif act == a_open_folder:
            self.open_attachment_folder()
        elif act == a_copy_path:
            _set_clipboard_text(ptxt)
        elif act == a_copy_folder:
            _set_clipboard_text(ftxt)
        elif act == a_copy_sel:
            try:
                copy_selected_cells(self.tbl)
            except Exception:
                pass
        elif act == a_del:
            self.delete_attachment_ui()

    # -------------------- preview find --------------------

    def _sync_find_enabled(self) -> None:
        try:
            is_text_tab = (self.preview_stack.currentIndex() == 2)
        except Exception:
            is_text_tab = False

        if not is_text_tab:
            try:
                self.btn_find.setEnabled(False)
                self.btn_next.setEnabled(False)
                self.btn_prev.setEnabled(False)
            except Exception:
                pass
            return

        try:
            has_text = bool((self.preview_txt.toPlainText() or "").strip())
        except Exception:
            has_text = False

        try:
            q = (self.ed_find.text() or "").strip()
        except Exception:
            q = ""

        ok = bool(has_text and q)
        try:
            self.btn_find.setEnabled(ok)
            self.btn_next.setEnabled(ok)
            self.btn_prev.setEnabled(ok)
        except Exception:
            pass

        if not ok:
            self._last_find_cursor = None
            self._last_find_text = ""
            try:
                self.lb_find.setText("")
            except Exception:
                pass

    def _find_first(self) -> None:
        self._last_find_cursor = None
        self._last_find_text = ""
        self._find_next()

    def _find_next(self) -> None:
        try:
            needle = (self.ed_find.text() or "").strip()
        except Exception:
            needle = ""
        if not needle:
            return

        try:
            doc = self.preview_txt.document()
        except Exception:
            return

        if self._last_find_text != needle:
            self._last_find_text = needle
            self._last_find_cursor = None

        try:
            if self._last_find_cursor is None:
                cur = QTextCursor(doc)
                cur.movePosition(QTextCursor.Start)
            else:
                cur = QTextCursor(self._last_find_cursor)
                cur.movePosition(QTextCursor.Right)

            found = doc.find(needle, cur)
            if found is not None and (not found.isNull()):
                self.preview_txt.setTextCursor(found)
                self._last_find_cursor = found
                try:
                    self.lb_find.setText("Nađeno")
                except Exception:
                    pass
                return

            try:
                self.lb_find.setText("Nema više (kraj) — vraćam na početak")
            except Exception:
                pass

            cur2 = QTextCursor(doc)
            cur2.movePosition(QTextCursor.Start)
            found2 = doc.find(needle, cur2)
            if found2 is not None and (not found2.isNull()):
                self.preview_txt.setTextCursor(found2)
                self._last_find_cursor = found2
                try:
                    self.lb_find.setText("Nađeno (od početka)")
                except Exception:
                    pass
            else:
                try:
                    self.lb_find.setText("Nije nađeno")
                except Exception:
                    pass
        except Exception:
            try:
                self.lb_find.setText("Greška u pretrazi")
            except Exception:
                pass

    def _find_prev(self) -> None:
        try:
            needle = (self.ed_find.text() or "").strip()
        except Exception:
            needle = ""
        if not needle:
            return

        try:
            doc = self.preview_txt.document()
            cur_now = self.preview_txt.textCursor()
            cur = QTextCursor(cur_now)
            cur.movePosition(QTextCursor.Left)
        except Exception:
            return

        try:
            found = doc.find(needle, cur, QTextDocument.FindBackward)
            if found is not None and (not found.isNull()):
                self.preview_txt.setTextCursor(found)
                self._last_find_cursor = found
                self._last_find_text = needle
                try:
                    self.lb_find.setText("Nađeno")
                except Exception:
                    pass
                return

            cur2 = QTextCursor(doc)
            cur2.movePosition(QTextCursor.End)
            found2 = doc.find(needle, cur2, QTextDocument.FindBackward)
            if found2 is not None and (not found2.isNull()):
                self.preview_txt.setTextCursor(found2)
                self._last_find_cursor = found2
                self._last_find_text = needle
                try:
                    self.lb_find.setText("Nađeno (od kraja)")
                except Exception:
                    pass
            else:
                try:
                    self.lb_find.setText("Nije nađeno")
                except Exception:
                    pass
        except Exception:
            try:
                self.lb_find.setText("Greška u pretrazi")
            except Exception:
                pass

    # -------------------- keyboard shortcuts --------------------

    def keyPressEvent(self, event) -> None:  # type: ignore[override]
        try:
            if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_F:
                if self.tabs.currentWidget() == self.attach_tab:
                    try:
                        self.preview_stack.setCurrentIndex(2)
                    except Exception:
                        pass
                    try:
                        self.ed_find.setFocus()
                        self.ed_find.selectAll()
                    except Exception:
                        pass
                    event.accept()
                    return
        except Exception:
            pass
        super().keyPressEvent(event)

# (FILENAME: ui/asset_detail_dialog.py - END PART 4/4)